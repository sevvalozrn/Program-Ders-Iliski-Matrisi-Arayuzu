from tkinter import *
from PIL import Image, ImageTk
from tkinter import ttk, messagebox
import pyodbc
import pandas as pd
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.comments import Comment


def create_connection():
    """Veritabanı bağlantısını oluşturur ve geri döndürür"""
    conn = pyodbc.connect('DRIVER={ODBC Driver 17 for SQL Server};'
                          'SERVER=ASUS\\SQLEXPRESS;'
                          'DATABASE=RelationMatrix;'
                          'Trusted_Connection=yes')
    return conn


def check_database():
    conn = get_connection("master")
    conn.autocommit = True

    cursor = conn.cursor()

    cursor.execute(f"SELECT database_id FROM sys.databases WHERE Name = 'RelationMatrix'")
    database_exists = cursor.fetchone()

    if database_exists:
        print("There is already a database called RelationMatrix")
    else:
        cursor.execute("CREATE DATABASE RelationMatrix")
        print("RelationMatrix database created")

    conn.close()


def check_tables():
    conn = get_connection("RelationMatrix")
    conn.autocommit = True
    cursor = conn.cursor()

    cursor.execute("""
            IF NOT EXISTS (SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = 'Students')
            CREATE TABLE Students (
                Student INT,
                lesson_id INT,
                PRIMARY KEY (Student, lesson_id)
            );
        """)
    print("Students table checked/created.")

    cursor.execute("""
        IF NOT EXISTS (SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = 'Lessons')
        CREATE TABLE Lessons (
            id INT PRIMARY KEY IDENTITY(1,1),
            name NVARCHAR(255) NOT NULL
        );
    """)
    print("Lessons table checked/created.")

    tables = ["CourseOutcomes", "ProgramOutcomes", "ProgramCourseRelations", "EvaluationCriteria",
              "CourseEvaluationRelations"]
    for table in tables:
        cursor.execute(f"SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = '{table}'")
        table_exists = cursor.fetchone()[0]
        if table_exists:
            print(f"{table} table already exists.")
        else:
            if table == 'CourseOutcomes':
                cursor.execute('''
                    CREATE TABLE CourseOutcomes (
                        id INT PRIMARY KEY IDENTITY(1,1),
                        data TEXT NOT NULL,
                        LessonID INT NOT NULL,
                        FOREIGN KEY (LessonID) REFERENCES Lessons(id)
                    );
                ''')
            elif table == 'ProgramOutcomes':
                cursor.execute('''
                    CREATE TABLE ProgramOutcomes (
                        id INT PRIMARY KEY IDENTITY(1,1),
                        data TEXT NOT NULL,
                        LessonID INT NOT NULL,
                        FOREIGN KEY (LessonID) REFERENCES Lessons(id)
                    );
                ''')
            elif table == 'ProgramCourseRelations':
                cursor.execute('''
                    CREATE TABLE ProgramCourseRelations (
                        ProgramOutcomeID INT NOT NULL,
                        CourseOutcomeID INT NOT NULL,
                        RelationValue FLOAT NOT NULL CHECK (RelationValue BETWEEN 0 AND 1),
                        LessonID INT NOT NULL,
                        PRIMARY KEY (ProgramOutcomeID, CourseOutcomeID, LessonID),
                        FOREIGN KEY (ProgramOutcomeID) REFERENCES ProgramOutcomes(id),
                        FOREIGN KEY (CourseOutcomeID) REFERENCES CourseOutcomes(id),
                        FOREIGN KEY (LessonID) REFERENCES Lessons(id)
                    );
                ''')
            elif table == 'EvaluationCriteria':
                cursor.execute('''
                    CREATE TABLE EvaluationCriteria (
                        Criteria NVARCHAR(25) NOT NULL,
                        Weight INT NOT NULL,
                        LessonID INT NOT NULL,
                        PRIMARY KEY (Criteria, LessonID),
                        FOREIGN KEY (LessonID) REFERENCES Lessons(id)
                    );
                ''')
            elif table == 'CourseEvaluationRelations':
                cursor.execute('''
                    CREATE TABLE CourseEvaluationRelations (
                        CourseOutcomeID INT NOT NULL,
                        Criteria NVARCHAR(25),
                        RelationValue INT NOT NULL CHECK (RelationValue IN (0, 1)),
                        LessonID INT NOT NULL,
                        PRIMARY KEY (CourseOutcomeID, Criteria, LessonID),
                        FOREIGN KEY (CourseOutcomeID) REFERENCES CourseOutcomes(id),
                        FOREIGN KEY (Criteria, LessonID) REFERENCES EvaluationCriteria(Criteria, LessonID),
                        FOREIGN KEY (LessonID) REFERENCES Lessons(id)
                    );
                ''')
            print(f"{table} table created.")

    conn.close()


def create_table3():
    workbook = Workbook()

    # Ders ID'lerine göre veri filtreleme işlemi
    lesson_names = fetch_lesson_names()
    course_evaluation_relations = fetch_evaluation_relations()
    course_outcomes = fetch_table_data("CourseOutcomes")
    evaluation_criteria = fetch_evaluation_data()

    # EvaluationCriteria verilerinin LessonID bazında gruplanması
    criteria_weights = {
        (lesson_id, criteria): weight
        for criteria, weight, lesson_id in evaluation_criteria
    }

    for lesson_id, lesson_name in lesson_names.items():
        sheet = workbook.create_sheet(lesson_name)

        sheet.merge_cells('A1:B1')
        sheet['A1'] = f"Table 3 - {lesson_name}"
        sheet['C1'] = "Weighted Evaluation"

        sheet.merge_cells('A2:B2')
        sheet['A2'] = "Course Outcomes"

        filtered_criteria = [
            criteria for _, criteria, _, relation_lesson_id in course_evaluation_relations
            if relation_lesson_id == lesson_id and (relation_lesson_id, criteria) in criteria_weights
        ]

        for col, criteria in enumerate(set(filtered_criteria), start=3):
            sheet.cell(row=2, column=col, value=criteria)

        total_col = len(set(filtered_criteria)) + 3
        sheet.cell(row=2, column=total_col, value="Total")

        weighted_data = {}
        filtered_course_evaluation_relations = [
            (course_outcome_id, criteria, relation_value)
            for course_outcome_id, criteria, relation_value, relation_lesson_id in course_evaluation_relations
            if relation_lesson_id == lesson_id
        ]

        filtered_course_outcomes = [
            (course_outcome_id, program_text)
            for course_outcome_id, program_text, course_lesson_id in course_outcomes
            if course_lesson_id == lesson_id
        ]

        for course_outcome_id, criteria, relation_value in filtered_course_evaluation_relations:
            weight = criteria_weights.get((lesson_id, criteria), 0)
            weighted_value = (relation_value * weight) / 100

            if course_outcome_id not in weighted_data:
                weighted_data[course_outcome_id] = {}

            weighted_data[course_outcome_id][criteria] = weighted_value

        for row_idx, (course_outcome_id, program_text) in enumerate(filtered_course_outcomes, start=3):
            sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            cell = sheet.cell(row=row_idx, column=1, value=course_outcome_id)
            comment = Comment(program_text, "Database")
            cell.comment = comment

            total = 0
            # Her kriterin değerinin eklenmesi
            for col_idx, criteria in enumerate(set(filtered_criteria), start=3):
                value = weighted_data.get(course_outcome_id, {}).get(criteria, 0)
                sheet.cell(row=row_idx, column=col_idx, value=value)
                total += value

            sheet.cell(row=row_idx, column=total_col, value=total)

    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    workbook.save("table3.xlsx")


def create_table4():
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    # Ders kriter ağırlıklarının çekilmesi
    cursor.execute("SELECT Criteria, Weight, LessonID FROM EvaluationCriteria;")
    evaluation_weights = {}
    for criteria, weight, lesson_id in cursor.fetchall():
        if lesson_id not in evaluation_weights:
            evaluation_weights[lesson_id] = {}
        evaluation_weights[lesson_id][criteria] = weight

    # Öğrenci verilerinin çekilmesi
    cursor.execute("SELECT * FROM Students;")
    student_columns = [column[0] for column in cursor.description]
    student_data = cursor.fetchall()

    # Ders çıktıları (CourseOutcomes) bilgilerinin çekilmesi
    cursor.execute("SELECT id, data, LessonID FROM CourseOutcomes;")
    course_outcomes = {}
    outcome_descriptions = {}
    for outcome_id, data, lesson_id in cursor.fetchall():
        if lesson_id not in course_outcomes:
            course_outcomes[lesson_id] = []
        course_outcomes[lesson_id].append((outcome_id, data))
        outcome_descriptions[outcome_id] = data

    # Ders değerlendirme ilişkilerinin çekilmesi
    cursor.execute("SELECT CourseOutcomeID, Criteria, RelationValue, LessonID FROM CourseEvaluationRelations;")
    course_evaluation_relations = {}
    for outcome_id, criteria, relation_value, lesson_id in cursor.fetchall():
        if lesson_id not in course_evaluation_relations:
            course_evaluation_relations[lesson_id] = {}
        if outcome_id not in course_evaluation_relations[lesson_id]:
            course_evaluation_relations[lesson_id][outcome_id] = {}
        course_evaluation_relations[lesson_id][outcome_id][criteria] = relation_value

    cursor.execute("SELECT lesson_id, course_outcome_id, total_score FROM Table3;")
    total_scores = {}
    for lesson_id, course_outcome_id, total_score in cursor.fetchall():
        if lesson_id not in total_scores:
            total_scores[lesson_id] = {}
        total_scores[lesson_id][course_outcome_id] = total_score

    conn.close()

    lesson_names = fetch_lesson_names()

    workbook = Workbook()
    sheet_created = {}

    for student in student_data:
        student_lesson_id = student[student_columns.index("lesson_id")]
        student_id = student[student_columns.index("Student")]
        lesson_name = lesson_names.get(student_lesson_id, f"Lesson {student_lesson_id}")

        if student_lesson_id not in sheet_created:
            sheet = workbook.create_sheet(title=lesson_name)
            sheet_created[student_lesson_id] = sheet

            sheet.merge_cells("A1:B1")
            sheet["A1"] = f"{lesson_name} - Table 4"

            sheet["A2"] = "Student ID"
            sheet["B2"] = "Course Outcomes"

            # Kriterler, toplam, maksimum başarı oranı başlıklarının oluşturulması
            headers = [*evaluation_weights[student_lesson_id].keys(), "Total", "Max", "% Success"]
            for col_index, header in enumerate(headers, start=3):
                sheet.cell(row=2, column=col_index, value=header)

        sheet = sheet_created[student_lesson_id]
        current_row = sheet.max_row + 1

        for outcome_id, outcome_text in course_outcomes[student_lesson_id]:
            sheet.cell(row=current_row, column=1, value=student_id)
            sheet.cell(row=current_row, column=2, value=outcome_text)

            sheet.cell(row=current_row, column=2).comment = Comment(outcome_descriptions[outcome_id], "System")

            row = []
            total = 0

            for criteria, weight in evaluation_weights[student_lesson_id].items():
                score = student[student_columns.index(criteria)]
                relation_value = course_evaluation_relations[student_lesson_id].get(outcome_id, {}).get(criteria, 0)

                if score is not None and relation_value != 0:

                    weighted_score = score * weight * relation_value / 100
                else:
                    weighted_score = 0
                row.append(weighted_score)
                total += weighted_score

            max_score = total_scores.get(student_lesson_id, {}).get(outcome_id, 0) * 100
            success_rate = (total / max_score * 100) if max_score > 0 else 0

            row.extend([total, max_score, round(success_rate, 1)])

            for col_index, value in enumerate(row, start=3):
                sheet.cell(row=current_row, column=col_index, value=value)

            current_row += 1

    if 'Sheet' in workbook.sheetnames:
        if len(workbook.sheetnames) == 1:
            # Yeni bir sayfa eklemeden önce varsayılan sayfayı silmeyin
            workbook.create_sheet(title="DefaultSheet")
        del workbook['Sheet']
    workbook.save("table4.xlsx")


def create_table5():
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    cursor.execute("SELECT id, name FROM Lessons;")
    lessons = {row[0]: row[1] for row in cursor.fetchall()}

    cursor.execute("SELECT * FROM Students;")
    student_columns = [column[0] for column in cursor.description]
    student_data = cursor.fetchall()

    cursor.execute("SELECT id, data, LessonID FROM ProgramOutcomes;")
    program_outcomes = {}
    for program_id, data, lesson_id in cursor.fetchall():
        if lesson_id not in program_outcomes:
            program_outcomes[lesson_id] = {}
        program_outcomes[lesson_id][program_id] = data

    cursor.execute("SELECT id, data, LessonID FROM CourseOutcomes;")
    course_outcomes = {}
    for outcome_id, data, lesson_id in cursor.fetchall():
        if lesson_id not in course_outcomes:
            course_outcomes[lesson_id] = []
        course_outcomes[lesson_id].append((outcome_id, data))

    cursor.execute("SELECT ProgramOutcomeID, CourseOutcomeID, RelationValue, LessonID FROM ProgramCourseRelations;")
    relations = {}
    for program_id, course_id, relation_value, lesson_id in cursor.fetchall():
        if lesson_id not in relations:
            relations[lesson_id] = {}
        if program_id not in relations[lesson_id]:
            relations[lesson_id][program_id] = {}
        relations[lesson_id][program_id][course_id] = relation_value

    cursor.execute("SELECT student_id, course_outcome_id, success_rate FROM Table4;")
    success_rates = {}
    for student_id, course_outcome_id, success_rate in cursor.fetchall():
        if student_id not in success_rates:
            success_rates[student_id] = {}
        success_rates[student_id][course_outcome_id] = success_rate

    conn.close()
    workbook = Workbook()

    for lesson_id, lesson_name in lessons.items():
        sheet = workbook.create_sheet(title=f"{lesson_name}")
        sheet.merge_cells("A1:C1")
        sheet["A1"] = f"{lesson_name} - Table 5"

        sheet["A2"] = "Student ID"
        sheet["B2"] = "Program Outcomes"
        lesson_course_outcomes = course_outcomes.get(lesson_id, [])
        for idx, (_, course_outcome_text) in enumerate(lesson_course_outcomes, start=3):
            sheet.cell(row=2, column=idx, value=course_outcome_text)

        sheet.cell(row=2, column=len(lesson_course_outcomes) + 3, value="Success Rate")

        current_row = 3
        for student in student_data:
            student_id = student[student_columns.index("Student")]
            student_lesson_id = student[student_columns.index("lesson_id")]

            if student_lesson_id != lesson_id:
                continue

            lesson_program_outcomes = program_outcomes.get(lesson_id, {})

            for program_outcome_id, program_outcome_text in lesson_program_outcomes.items():
                relation_values = []
                row_values = []

                for course_outcome_id, _ in lesson_course_outcomes:
                    relation_value = relations.get(lesson_id, {}).get(program_outcome_id, {}).get(course_outcome_id, 0)
                    success_rate = success_rates.get(student_id, {}).get(course_outcome_id, 0)
                    weighted_value = relation_value * success_rate
                    relation_values.append(relation_value)
                    row_values.append(weighted_value)

                total_success = sum(row_values)
                num_course_outcomes = len(lesson_course_outcomes)
                avg_success = total_success / num_course_outcomes if num_course_outcomes else 0
                avg_relation_value = sum(relation_values) / len(relation_values) if relation_values else 0
                ratio = avg_success / avg_relation_value if avg_relation_value else 0

                sheet.cell(row=current_row, column=1, value=student_id)
                sheet.cell(row=current_row, column=2, value=program_outcome_text)
                for col_index, value in enumerate(row_values, start=3):
                    sheet.cell(row=current_row, column=col_index, value=round(value, 1))
                sheet.cell(row=current_row, column=len(lesson_course_outcomes) + 3, value=round(ratio, 1))

                current_row += 1

    del workbook["Sheet"]
    workbook.save("table5.xlsx")


def create_notes():
    workbook = Workbook()
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    # Ders adlarını almak için fetch_lesson_names fonksiyonu kullanımı
    lesson_names = fetch_lesson_names()

    # Öğrenciler tablosundaki derslerin alınması
    cursor.execute("SELECT DISTINCT lesson_id FROM Students;")
    lesson_ids = [row[0] for row in cursor.fetchall()]

    if not lesson_ids:
        print("No lessons found in Students table.")
        return

    for lesson_id in lesson_ids:
        lesson_name = lesson_names.get(lesson_id, "Unknown Lesson")

        sheet = workbook.create_sheet(title=f"Lesson {lesson_id}")
        sheet.merge_cells('A1:B1')
        sheet['A1'] = f"Table Note - {lesson_name}"
        sheet['C1'] = "Notes"

        cursor.execute("""
            SELECT ec.Criteria, ec.Weight
            FROM EvaluationCriteria ec
            WHERE ec.LessonID = ?;
        """, (lesson_id,))
        criteria_weights = {row[0]: row[1] for row in cursor.fetchall()}

        if not criteria_weights:
            print(f"No criteria found for Lesson ID {lesson_id}.")
            continue

        criteria_columns = [f"[{criterion}]" for criterion in criteria_weights.keys()]
        columns_str = ", ".join(criteria_columns)

        cursor.execute(f"""
            SELECT s.Student, s.lesson_id, {columns_str}
            FROM Students s
            WHERE s.lesson_id = ?;
        """, (lesson_id,))
        rows = cursor.fetchall()

        if not rows:
            print(f"No data found for Lesson ID {lesson_id}.")
            continue

        columns = ["Student "] + list(criteria_weights.keys()) + ["Average"]
        for col_idx, column_name in enumerate(columns, start=1):
            sheet.cell(row=2, column=col_idx, value=column_name)

        for row_idx, row in enumerate(rows, start=3):
            student_id = row[0]
            sheet.cell(row=row_idx, column=1, value=student_id)

            total_score = 0
            weight_sum = 0
            for col_idx, (criterion, value) in enumerate(zip(criteria_weights.keys(), row[2:]), start=2):
                if value is not None:
                    sheet.cell(row=row_idx, column=col_idx, value=value)
                    weight = criteria_weights.get(criterion, 0)
                    total_score += value * weight
                    weight_sum += weight
                else:
                    sheet.cell(row=row_idx, column=col_idx, value=0)

            average = total_score / 100 if weight_sum == 100 else total_score / weight_sum
            sheet.cell(row=row_idx, column=len(columns), value=round(average, 2))

    del workbook['Sheet']
    workbook.save("notlar.xlsx")
    conn.close()


def get_connection(database=None):
    connection_string = (
        "Driver={ODBC Driver 17 for SQL Server};"
        "Server=ASUS\\SQLEXPRESS;"
        f"Database={database or 'master'};"
        "Trusted_Connection=yes;"
    )
    return pyodbc.connect(connection_string)


def insert_data_into_table(table_name, data, lesson_id):
    conn = get_connection("RelationMatrix")
    conn.autocommit = True
    cursor = conn.cursor()

    if table_name == 'CourseOutcomes':
        cursor.execute("INSERT INTO CourseOutcomes (data, LessonID) VALUES (?, ?)", (data, lesson_id))
    elif table_name == 'ProgramOutcomes':
        cursor.execute("INSERT INTO ProgramOutcomes (data, LessonID) VALUES (?, ?)", (data, lesson_id))

    conn.close()


def delete_data_from_table_by_id(table_name, record_id):
    conn = get_connection("RelationMatrix")
    conn.autocommit = True
    cursor = conn.cursor()
    print("id", record_id)

    if table_name == 'CourseOutcomes':
        cursor.execute("DELETE FROM CourseOutcomes WHERE id = ?", (record_id,))
    elif table_name == 'ProgramOutcomes':
        cursor.execute("DELETE FROM ProgramOutcomes WHERE id = ?", (record_id,))
    conn.close()


def get_data_from_table(table_name):
    """Veritabanından belirtilen tablodan veri çeker"""
    conn = create_connection()
    cursor = conn.cursor()
    query = f"SELECT * FROM {table_name}"
    cursor.execute(query)
    data = cursor.fetchall()
    conn.close()
    return data, cursor.description


def display_data_in_treeview(table_name, frame, x, y, l_id):
    for widget in frame.winfo_children():
        if isinstance(widget, ttk.Treeview):
            widget.destroy()

    data, description = get_data_from_table_with_filter(table_name, l_id)

    columns = [desc[0] for desc in description]

    if table_name == 'Students':
        if "Student" in columns:
            columns = ["Student"]
            data = [[row[columns.index("Student")]] for row in data]

    treeview = ttk.Treeview(frame, columns=columns, show="headings")

    for column in columns:
        treeview.heading(column, text=column)
        treeview.column(column, stretch=True)

    max_lengths = {column: len(column) for column in columns}

    for row in data:
        row = [str(cell).strip() for cell in row]
        for i, value in enumerate(row):
            column = columns[i]
            max_lengths[column] = max(max_lengths[column], len(value))
        treeview.insert("", "end", values=row)

    for column in columns:
        treeview.column(column, width=max_lengths[column] * 10)

    row_height = 20
    total_height = (len(data) + 3) * row_height
    treeview.place(x=x, y=y, height=total_height)

    if table_name == 'Students' and "Student" in columns:
        student_no_index = columns.index("Student")

        if student_no_index is not None:
            treeview.bind(
                "<ButtonRelease-1>",
                lambda event: on_row_select(event, frame, treeview, student_no_index, l_id)
            )


def on_row_select(event, frame, tree, student_no, l_id):
    conn = get_connection("RelationMatrix")
    conn.autocommit = True
    cursor = conn.cursor()

    query = "SELECT name FROM Lessons WHERE id = ?"
    cursor.execute(query, (l_id,))
    result = cursor.fetchone()

    if result:
        lesson_name = result[0]
        print("Lesson Name:", lesson_name)
    else:
        print("Lesson not found for the selected_id:", l_id)

    cursor.close()
    conn.close()

    """Bir satır seçildiğinde ilgili verileri gösterir"""
    selected_item = tree.selection()[0]
    row_index = tree.index(selected_item)

    selected_item = tree.selection()
    if selected_item:
        selected_item_id = selected_item[0]
        row_values = tree.item(selected_item_id, "values")

        if student_no is not None:
            student_no = row_values[student_no]
            print(f"Selected Student No: {student_no}")
        else:
            print("Error: Student No index is None")

    file_path1 = "table4.xlsx"
    file_path2 = "table5.xlsx"

    wb = load_workbook(file_path1, data_only=True)
    sheet_names = wb.sheetnames

    if lesson_name in sheet_names:
        sheet1 = wb[lesson_name]

    df1 = pd.read_excel(file_path1, sheet_name=lesson_name, header=None)
    df1 = df1.dropna(how="all", axis=0)
    df1 = df1.dropna(how="all", axis=1)

    df1 = df1.fillna("")

    df1_reset = df1.reset_index(drop=True)

    if not df1.empty:
        df1_filtered = df1_reset.iloc[2:]
        df1_filtered = df1_filtered[df1_filtered.iloc[:, 0].astype(str) == str(student_no)]
    else:
        df1_filtered = pd.DataFrame()

    if not df1.empty:
        non_empty_values1 = [str(value) for value in df1.iloc[0].values if pd.notnull(value)]
        header_label_text1 = " ".join(non_empty_values1)
        header_label1 = Label(frame, text=header_label_text1, font=("Arial", 12, "bold"))
        header_label1.place(x=10, y=200)

    if len(df1) > 1:
        columns1 = list(df1.iloc[1].values)
        data1 = df1_filtered.reset_index(drop=True)  # Yeni index ile veriyi sıfırla
    else:
        columns1 = []
        data1 = pd.DataFrame()

    table4 = ttk.Treeview(frame, columns=columns1, show='headings', height=20)
    for col in columns1:
        table4.heading(col, text=col)
        table4.column(col, width=100)

    for _, row in data1.iterrows():
        table4.insert("", "end", values=list(row))

    row_height = 20
    total_height = (len(data1) + 4) * row_height
    table4.place(x=10, y=230, height=total_height)

    # İkinci tabloyu (table5) yükle
    df2 = pd.read_excel(file_path2, sheet_name=lesson_name, header=None)
    df2 = df2.dropna(how="all", axis=0)
    df2 = df2.dropna(how="all", axis=1)

    # Boş hücreleri 0 ile doldur
    df2 = df2.fillna("")

    # Dataframe index'ini sıfırlıyoruz
    df2_reset = df2.reset_index(drop=True)

    # Öğrenci numarasına göre filtreleme, 3. satırdan itibaren
    if not df2.empty:
        # 3. satırdan itibaren ilk sütun ile öğrenci numarasını karşılaştır
        df2_filtered = df2_reset.iloc[2:]  # 3. satırdan sonrasını al
        df2_filtered = df2_filtered[
            df2_filtered.iloc[:, 0].astype(str) == str(student_no)]  # İlk sütun öğrenci numarası
    else:
        df2_filtered = pd.DataFrame()

    if not df2.empty:
        non_empty_values2 = [str(value) for value in df2.iloc[0].values if pd.notnull(value)]
        header_label_text2 = " ".join(non_empty_values2)
        header_label2 = Label(frame, text=header_label_text2, font=("Arial", 12, "bold"))
        header_label2.place(x=10, y=470)

    # Sütun başlıklarını ikinci satırdan al
    if len(df2) > 1:
        columns2 = list(df2.iloc[1].values)
        data2 = df2_filtered.reset_index(drop=True)  # Yeni index ile veriyi sıfırla
    else:
        columns2 = []
        data2 = pd.DataFrame()

    # İkinci tabloyu (table5) oluştur
    table5 = ttk.Treeview(frame, columns=columns2, show='headings', height=20)
    for col in columns2:
        table5.heading(col, text=col)
        table5.column(col, width=100)

    # Satırları ekle
    for _, row in data2.iterrows():
        table5.insert("", "end", values=list(row))

    row_height = 20
    total_height2 = (len(data2) + 4) * row_height
    table5.place(x=10, y=500, height=total_height2)


def validate_input(value):
    try:
        numeric_value = float(value)
        if 0 <= numeric_value <= 1:
            return True
        else:
            messagebox.showerror("Geçersiz Değer", "Değer 0 ile 1 arasında olmalıdır.")
            return False
    except ValueError:
        messagebox.showerror("Geçersiz Giriş", "Lütfen bir sayı giriniz.")
        return False

# İliski Ekleme
def insert_relation_value(program_outcome_id, course_outcome_id, relation_value, lesson_id):

    if not validate_input(relation_value):
        return

    conn = get_connection("RelationMatrix")

def get_data_from_table_with_filter(table_name, l_id):
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()
    l_id = int(l_id)

    if table_name == 'Students':
        try:
            # LessonID'ye göre filtrelenmiş verileri çek
            query = f"SELECT * FROM {table_name} WHERE lesson_id = ?"
            cursor.execute(query, (l_id,))
            data = cursor.fetchall()
            description = cursor.description
        finally:
            conn.close()

        return data, description

    else:
        try:
            # LessonID'ye göre filtrelenmiş verileri çek
            query = f"SELECT * FROM {table_name} WHERE LessonID = ?"
            cursor.execute(query, (l_id,))
            data = cursor.fetchall()
            description = cursor.description
        finally:
            conn.close()

        return data, description


root = Tk()
root.title("KOCAELİ SAĞLIK VE TEKNOLOJİ ÜNİVERSİTESİ - Ders Verileri Giriş Ekranı")
root.geometry("1500x790+0+0")

frame1 = Frame(root)
frame2 = Frame(root)

root.grid_rowconfigure(0, weight=1)
root.grid_columnconfigure(0, weight=1)

for frame in (frame1, frame2):
    frame.grid(row=0, column=0, sticky="nsew")

# === Frame 1: Ders Seçim Ekranı ===
image = Image.open("kostu_yuvarlak_logo.png")
image = image.resize((200, 200), Image.Resampling.LANCZOS)
photo = ImageTk.PhotoImage(image)

image_label = Label(frame1, image=photo)
image_label.photo = photo
image_label.pack(pady=10)

label = Label(frame1, text="KOCAELİ SAĞLIK VE TEKNOLOJİ \n ÜNİVERSİTESİ", fg="green",
              font=('Times New Roman', 30, "bold"))
label.pack(padx=30, pady=10)

title_label = Label(frame1, text="Ders Seçimi Yapınız", font=("Arial", 16, "bold"))
title_label.pack(pady=10)


# Dropdown menü

def check_lessons():
    conn = get_connection("RelationMatrix")
    conn.autocommit = True
    cursor = conn.cursor()

    cursor.execute("""
        IF NOT EXISTS (SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = 'Lessons')
        CREATE TABLE Lessons (
            id INT PRIMARY KEY IDENTITY(1,1),
            name NVARCHAR(255) NOT NULL
        );
    """)

    cursor.execute("SELECT id, name FROM Lessons")
    courses = [f"{row[0]} - {row[1]}" for row in cursor.fetchall()]

    conn.close()
    return courses


selected_course = StringVar()
courses = check_lessons()
selected_course.set("0 - Ders Seçiniz")

selected_value = selected_course.get()
selected_id = selected_value.split(" - ")[0]
print(selected_id)

dropdown = ttk.Combobox(frame1, textvariable=selected_course, values=courses, state="readonly", width=40)
dropdown.pack(pady=10)

# Yeni ders ekleme
add_course_label = Label(frame1, text="Seçeceğiniz ders mevcut değilse, aşağıdan ekleyebilirsiniz:", font=("Arial", 10))
add_course_label.pack(pady=5)

new_course_entry = Entry(frame1, width=30)
new_course_entry.pack(pady=5)


def add_course():
    conn = get_connection("RelationMatrix")
    conn.autocommit = True
    cursor = conn.cursor()

    new_course = new_course_entry.get()
    if new_course:
        cursor.execute("INSERT INTO Lessons (name) VALUES (?)", (new_course))

        courses = check_lessons()
        dropdown['values'] = courses

        selected_course.set(f"{cursor.id} - {new_course}")

        new_course_entry.delete(0, END)

    else:
        status_label.config(text="Lütfen bir ders adı girin.", fg="red")


add_button = Button(frame1, text="Ders Ekle", command=add_course)
add_button.pack(pady=5)

status_label = Label(frame1, text="", font=("Arial", 10))
status_label.pack(pady=5)


def proceed_to_next():
    if selected_course.get() != "Ders Seçiniz":
        label = (f"Seçilen Ders: {selected_course.get()}")
        selected_value = selected_course.get()
        selected_id = selected_value.split(" - ")[0]
        print(selected_id)
        show_frame2(selected_id, label)
    else:
        status_label.config(text="Lütfen bir ders seçiniz.", fg="red")


next_button = Button(frame1, text="Devam Et", command=proceed_to_next)
next_button.pack(pady=10)


def insert_table(table_name, data, lesson_id, frame):
    insert_data_into_table(table_name, data, lesson_id)
    print(f"Data has been inserted into {table_name}.")
    display_data_in_treeview(table_name, frame, 10, 40, lesson_id)


def del_from_table(table_name, id, lesson_id, frame):
    delete_data_from_table_by_id(table_name, id)
    print(f"Data has been deleted from {table_name}, {id}.")
    display_data_in_treeview(table_name, frame, 10, 40, lesson_id)


def show_frame1():
    frame1.tkraise()

def validate_input(value):
    try:
        numeric_value = float(value)
        if 0 <= numeric_value <= 1:
            return True
        else:
            messagebox.showerror("Geçersiz Değer", "Değer 0 ile 1 arasında olmalıdır.")
            return False
    except ValueError:
        messagebox.showerror("Geçersiz Giriş", "Lütfen bir sayı giriniz.")
        return False


# İliski Ekleme
def insert_relation_value(program_outcome_id, course_outcome_id, relation_value, lesson_id):
    conn = get_connection("RelationMatrix")
    conn.autocommit = True
    cursor = conn.cursor()
    if not validate_input(relation_value):
        return

    cursor.execute('''
        INSERT INTO ProgramCourseRelations (ProgramOutcomeID, CourseOutcomeID, RelationValue, LessonID)
        VALUES (?, ?, ?, ?);
    ''', program_outcome_id, course_outcome_id, relation_value, lesson_id)

    print(
        f"Relation between ProgramOutcome {program_outcome_id} and CourseOutcome {course_outcome_id} for Lesson {lesson_id} has been inserted.")
    conn.close()


def fetch_table_data(table_name):
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    query = f"SELECT id, data, LessonID FROM {table_name};"
    cursor.execute(query)

    data = cursor.fetchall()

    cursor.close()
    conn.close()

    return data


def fetch_lesson_names():
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    query = "SELECT id, name FROM Lessons;"
    cursor.execute(query)

    lesson_names = dict(cursor.fetchall())

    cursor.close()
    conn.close()

    return lesson_names


def fetch_evaluation_data():
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    query = "SELECT Criteria, Weight, LessonID FROM EvaluationCriteria;"
    cursor.execute(query)
    data = cursor.fetchall()

    cursor.close()
    conn.close()

    return data


def fetch_evaluation_relations():
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    query = """
    SELECT CourseOutcomeID, Criteria, RelationValue, LessonID
    FROM CourseEvaluationRelations;
    """
    cursor.execute(query)

    relations = cursor.fetchall()

    cursor.close()
    conn.close()

    return relations


def fetch_relations():
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    query = """
    SELECT ProgramOutcomeID, CourseOutcomeID, RelationValue, LessonID
    FROM ProgramCourseRelations;
    """
    cursor.execute(query)

    relations = cursor.fetchall()

    cursor.close()
    conn.close()

    return relations


def create_table1():
    workbook = Workbook()
    lesson_names = fetch_lesson_names()

    program_outcomes = fetch_table_data("ProgramOutcomes")
    course_outcomes = fetch_table_data("CourseOutcomes")

    program_row_count = len(program_outcomes)
    course_row_count = len(course_outcomes)

    for lesson_id, lesson_name in lesson_names.items():
        sheet = workbook.create_sheet(lesson_name)

        sheet.merge_cells('A1:B1')
        sheet['A1'] = f"Table 1 - {lesson_name}"
        sheet.merge_cells('A2:B2')
        sheet['A2'] = "Program Outcomes"
        sheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=course_row_count + 2)
        sheet['C1'] = "Course Outcomes"

        for i, (program_id, program_text, program_lesson_id) in enumerate(program_outcomes, start=1):
            if program_lesson_id == lesson_id:  # LessonID'ye göre filtreleme işlemi
                sheet.merge_cells(f"A{i + 2}:B{i + 2}")
                cell = sheet[f"A{i + 2}"]
                cell.value = program_id
                comment = Comment(program_text, "Database")
                cell.comment = comment

        for j, (course_id, course_text, course_lesson_id) in enumerate(course_outcomes, start=1):
            if course_lesson_id == lesson_id:  # LessonID'ye göre filtreleme işlemi
                c = sheet.cell(row=2, column=j + 2)
                c.value = course_id
                comment = Comment(course_text, "Database")
                c.comment = comment

        relations = fetch_relations()
        for relation in relations:
            program_outcome_id, course_outcome_id, relation_value, relation_lesson_id = relation
            if relation_lesson_id == lesson_id:
                row = program_outcome_id + 2
                col = course_outcome_id + 2
                sheet.cell(row=row, column=col, value=relation_value)

        for i in range(3, program_row_count + 3):
            course_outcomes_count = sum(
                1 for (course_id, _, course_lesson_id) in course_outcomes if course_lesson_id == lesson_id
            )

            total = sum(sheet.cell(row=i, column=j).value or 0 for j in range(3, course_row_count + 3))
            result = round((total / course_outcomes_count), 2) if course_outcomes_count > 0 else 0
            sheet.cell(row=i, column=course_row_count + 3, value=result)

        sheet.cell(row=2, column=course_row_count + 3, value="Rel Value")
    del workbook['Sheet']

    workbook.save(filename="table1.xlsx")


def create_table2():
    workbook = Workbook()
    lesson_names = fetch_lesson_names()

    for lesson_id, lesson_name in lesson_names.items():
        sheet = workbook.create_sheet(lesson_name)

        sheet.merge_cells('A1:B1')
        sheet['A1'] = f"Table 2 - {lesson_name}"
        sheet.merge_cells('A2:B2')
        sheet['A2'] = "Course Outcomes"

        course_outcomes = fetch_table_data("CourseOutcomes")
        filtered_course_outcomes = [
            (course_id, course_text) for course_id, course_text, course_lesson_id in course_outcomes
            if course_lesson_id == lesson_id
        ]
        for i, (course_id, course_text) in enumerate(filtered_course_outcomes, start=1):
            sheet.merge_cells(f"A{i + 2}:B{i + 2}")
            cell = sheet[f"A{i + 2}"]
            cell.value = course_id
            comment = Comment(course_text, "Database")
            cell.comment = comment

        evaluation_criteria = fetch_evaluation_data()
        filtered_criteria = [
            (criteria, weight) for criteria, weight, criteria_lesson_id in evaluation_criteria
            if criteria_lesson_id == lesson_id
        ]
        for index, (criteria, weight) in enumerate(filtered_criteria, start=3):
            sheet.cell(row=1, column=index, value=weight)
            sheet.cell(row=2, column=index, value=criteria)

        evaluation_relations = fetch_evaluation_relations()
        for relation in evaluation_relations:
            course_outcome_id, criteria, relation_value, relation_lesson_id = relation
            if relation_lesson_id == lesson_id:
                row = None
                for i, (course_id, _) in enumerate(filtered_course_outcomes, start=3):
                    if course_outcome_id == course_id:
                        row = i
                        break

                col = None
                for j, (crit, _) in enumerate(filtered_criteria, start=3):
                    if crit == criteria:
                        col = j
                        break

                if row and col:
                    sheet.cell(row=row, column=col, value=relation_value)

        total_col = len(filtered_criteria) + 3
        sheet.cell(row=2, column=total_col, value="Total")
        for row_idx in range(3, len(filtered_course_outcomes) + 3):
            total = sum(sheet.cell(row=row_idx, column=col_idx).value or 0 for col_idx in range(3, total_col))
            sheet.cell(row=row_idx, column=total_col, value=total)

    del workbook['Sheet']
    workbook.save(filename="table2.xlsx")


def show_excel(table, frame, lesson_name):
    create_table2()
    create_table1()

    file_path = table + ".xlsx"
    wb = load_workbook(file_path, data_only=True)
    sheet_names = wb.sheetnames

    if lesson_name in sheet_names:
        sheet = wb[lesson_name]

    df = pd.read_excel(file_path, sheet_name=lesson_name, header=None)

    df = df.dropna(how="all", axis=0)
    df = df.dropna(how="all", axis=1)
    df = df[df.iloc[:, 0].notna()]

    second_row = df.iloc[1]
    non_empty_columns = second_row.dropna().index.tolist()

    df = df[non_empty_columns]

    if len(df) > 1:
        columns = list(df.iloc[1].values)
        data = df.iloc[2:].reset_index(drop=True)
    else:
        columns = []
        data = pd.DataFrame()

    treeview = ttk.Treeview(frame, columns=columns, show='headings', height=10)

    if table == "table2":
        for idx, col in enumerate(columns):
            treeview.heading(col, text=col)

            if idx == 0:
                treeview.column(col, width=900)
            else:
                treeview.column(col, width=50)
        comments_dict = {}

        for row_num, row in enumerate(sheet.iter_rows()):
            for col_num, cell in enumerate(row):
                if cell.comment:
                    comments_dict[(row_num, col_num)] = cell.comment.text

        for row_index, row in data.iterrows():
            row_values = list(row)
            for col_index, value in enumerate(row_values):
                if pd.isna(value):
                    value = ""

                comment = comments_dict.get((row_index + 2, col_index))
                if comment:
                    value = f"{value}. {comment}"
                row_values[col_index] = value

            treeview.insert("", "end", values=row_values)

    else:
        for idx, col in enumerate(columns):
            treeview.heading(col, text=col)

            if idx == 0:
                treeview.column(col, width=100)
            else:
                treeview.column(col, width=100)

        for row_index, row in data.iterrows():
            row_values = list(row)

            for col_index, value in enumerate(row_values):
                if pd.isna(value):
                    value = ""
                row_values[col_index] = value
            treeview.insert("", "end", values=row_values)

    row_height = 20
    treeview.grid(row=2, column=0, padx=5, pady=5)


# === Frame 2: Seçilen Ders Ekranı ===

def show_frame2(l_id, text):
    frame2 = Frame(root)
    frame2.grid(row=0, column=0, sticky="nsew")

    selected_id = l_id
    selected_label = Label(frame2, text=text, font=("Arial", 14))
    selected_label.pack(pady=5)

    info_label = Label(frame2, text="Ders seçim ekranından çıktınız.", font=("Arial", 10))
    info_label.pack(pady=1)

    menu_frame = Frame(frame2, bg='#383838')
    menu_frame.pack(side=LEFT, fill=Y, padx=5, pady=5)
    menu_frame.pack_propagate(flag=False)
    menu_frame.configure(width=120)

    menu_button1 = Button(menu_frame, text="Ders Seçimine\nDön", bg='#383838', fg='white', bd=0,
                          activebackground='#383838',
                          command=lambda: show_frame1(), anchor='w')
    menu_button1.place(x=9, y=5, height=40)
    button1_ind = Label(menu_frame, bg='white')
    button1_ind.place(x=3, y=5, height=40, width=3)

    menu_button2 = Button(menu_frame, text="Ders Çıktıları", bg='#383838', fg='white', bd=0, activebackground='#383838',
                          command=lambda: show_other_frames(3, frame2, selected_id), anchor='w')
    menu_button2.place(x=9, y=50, height=40)
    button2_ind = Label(menu_frame, bg='white')
    button2_ind.place(x=3, y=50, height=40, width=3)

    menu_button3 = Button(menu_frame, text="Program Çıktıları", bg='#383838', fg='white', bd=0,
                          activebackground='#383838',
                          command=lambda: show_other_frames(4, frame2, selected_id), anchor='w')
    menu_button3.place(x=9, y=95, height=40)
    button3_ind = Label(menu_frame, bg='white')
    button3_ind.place(x=3, y=95, height=40, width=3)

    menu_button4 = Button(menu_frame, text="Program - Ders\nÇıktısı İşlemleri", bg='#383838', fg='white', bd=0,
                          activebackground='#383838', command=lambda: show_other_frames(5, frame2, selected_id),
                          anchor='w')
    menu_button4.place(x=9, y=140, height=40)
    button4_ind = Label(menu_frame, bg='white')
    button4_ind.place(x=3, y=140, height=40, width=3)

    menu_button5 = Button(menu_frame, text="Öğrenci Listesi", bg='#383838', fg='white', bd=0,
                          activebackground='#383838', command=lambda: show_other_frames(6, frame2, selected_id),
                          anchor='w')
    menu_button5.place(x=9, y=185, height=40)
    button5_ind = Label(menu_frame, bg='white')
    button5_ind.place(x=3, y=185, height=40, width=3)

    menu_button6 = Button(menu_frame, text="Değerlendirme\nkriterleri\nve ağırlıkları\ngiriş ekranı", bg='#383838',
                          fg='white', bd=0,
                          activebackground='#383838', command=lambda: show_other_frames(7, frame2, selected_id),
                          anchor='w')
    menu_button6.place(x=9, y=230, height=60)
    button6_ind = Label(menu_frame, bg='white')
    button6_ind.place(x=3, y=230, height=60, width=3)

    menu_button7 = Button(menu_frame, text="Öğrenci Notları\nGiriş Ekranı", bg='#383838',
                          fg='white', bd=0,
                          activebackground='#383838', command=lambda: show_other_frames(8, frame2, selected_id),
                          anchor='w')
    menu_button7.place(x=9, y=295, height=60)
    button7_ind = Label(menu_frame, bg='white')
    button7_ind.place(x=3, y=295, height=60, width=3)

    menu_button8 = Button(menu_frame, text="Ders Çıktısı ve\n Değerlendirme\n Kriteri Girişi", bg='#383838',
                          fg='white', bd=0,
                          activebackground='#383838', command=lambda: show_other_frames(9, frame2, selected_id),
                          anchor='w')
    menu_button8.place(x=9, y=360, height=60)
    button8_ind = Label(menu_frame, bg='white')
    button8_ind.place(x=3, y=360, height=60, width=3)


def show_other_frames(targetf, frame2, selected_id):
    conn = get_connection("RelationMatrix")
    conn.autocommit = True
    cursor = conn.cursor()

    query = "SELECT name FROM Lessons WHERE id = ?"
    cursor.execute(query, (selected_id,))
    result = cursor.fetchone()

    if result:
        lesson_name = result[0]  # 'name' sütunu ilk sırada olacaktır
        print("Lesson Name:", lesson_name)
    else:
        print("Lesson not found for the selected_id:", selected_id)

    cursor.close()
    conn.close()

    if not hasattr(show_other_frames, "active_frames"):
        show_other_frames.active_frames = []

    for frame in show_other_frames.active_frames:
        frame.destroy()
    show_other_frames.active_frames.clear()

    if targetf == 3:
        frame3 = Frame(frame2)
        show_other_frames.active_frames.append(frame3)
        frame3_title = Label(frame3, text="Ders Öğrenme Çıktıları", font=("Arial", 12))
        frame3_title.place(x=10, y=10)

        display_data_in_treeview("CourseOutcomes", frame3, 10, 40, selected_id)

        c_panel_frame = Frame(frame3)
        c_panel_frame.place(x=0, y=620)

        c_outcome_entry = Entry(c_panel_frame, width=40, font=("Arial", 10))
        c_outcome_entry.grid(row=1, column=0, padx=5, pady=5)

        co_info = Label(c_panel_frame,
                        text="Ders Çıktısı eklerken metin giriniz,\n silmek için seçilen Ders Çıktısının Id'sini giriniz.",
                        font=("Arial", 8))
        co_info.grid(row=0, column=0, padx=5, pady=5)

        add_button = Button(c_panel_frame, text="Ders Çıktısı Ekle", font=("Arial", 10),
                            command=lambda: insert_table("CourseOutcomes", c_outcome_entry.get(), selected_id, frame3))
        add_button.grid(row=1, column=1, padx=5, pady=5)

        delete_button = Button(c_panel_frame, text="Seçilen Çıktıyı Sil", font=("Arial", 10),
                               command=lambda: del_from_table("CourseOutcomes", c_outcome_entry.get(), selected_id,
                                                              frame3))
        delete_button.grid(row=1, column=2, padx=5, pady=5)

        frame3.pack(side=LEFT, fill=BOTH, expand=True, padx=5, pady=5)

    elif targetf == 4:
        frame4 = Frame(frame2)
        show_other_frames.active_frames.append(frame4)
        frame4_title = Label(frame4, text="Program Çıktısı İşlemleri", font=("Arial", 12))
        frame4_title.place(x=10, y=10)

        display_data_in_treeview("ProgramOutcomes", frame4, 10, 40, selected_id)

        p_panel_frame = Frame(frame4)
        p_panel_frame.place(x=0, y=620)

        p_outcome_entry = Entry(p_panel_frame, width=40, font=("Arial", 10))
        p_outcome_entry.grid(row=1, column=0, padx=5, pady=5)

        po_info = Label(p_panel_frame,
                        text="Program Çıktısı eklerken metin giriniz, silmek için seçilen Ders Çıktısının Id'sini giriniz.",
                        font=("Arial", 8))
        po_info.grid(row=0, column=0, padx=5, pady=5)

        add_button = Button(p_panel_frame, text="Program Çıktısı Ekle", font=("Arial", 10),
                            command=lambda: insert_table("ProgramOutcomes", p_outcome_entry.get(), selected_id, frame4))
        add_button.grid(row=1, column=1, padx=5, pady=5)

        delete_button = Button(p_panel_frame, text="Seçilen Çıktıyı Sil", font=("Arial", 10),
                               command=lambda: del_from_table("ProgramOutcomes", p_outcome_entry.get(), selected_id,
                                                              frame4))
        delete_button.grid(row=1, column=2, padx=5, pady=5)

        frame4.pack(side=LEFT, fill=BOTH, expand=True, padx=5, pady=5)

    elif targetf == 5:
        frame5 = Frame(frame2)
        show_other_frames.active_frames.append(frame5)
        frame5_title = Label(frame5, text="Ders - Program Çıktısı İşlemleri", font=("Arial", 12))
        frame5_title.place(x=0, y=0)

        kri_frame = Frame(frame5)
        kri_frame.place(x=0, y=30)

        frame = Frame(frame5)
        frame.place(x=100, y=400)

        show_button = Button(kri_frame, text="Tablo 1 Görüntüle", font=("Arial", 10),
                             command=lambda: show_excel("table1", kri_frame, lesson_name))
        show_button.grid(row=0, column=0, padx=5, pady=5)

        dk1_info = Label(frame, text="Ders ID'sini giriniz:", font=("Arial", 10))
        dk1_info.grid(row=1, column=0, padx=5, pady=5)

        dk1_entry = Entry(frame, width=40, font=("Arial", 10))
        dk1_entry.grid(row=1, column=1, padx=5, pady=5)

        dk2_info = Label(frame, text="Program Çıktısı ID'sini giriniz:", font=("Arial", 10))
        dk2_info.grid(row=2, column=0, padx=5, pady=5)

        dk2_entry = Entry(frame, width=40, font=("Arial", 10))
        dk2_entry.grid(row=2, column=1, padx=5, pady=5)

        dk3_info = Label(frame, text="Ders Çıktısı ID'sini giriniz:", font=("Arial", 10))
        dk3_info.grid(row=3, column=0, padx=5, pady=5)

        dk3_entry = Entry(frame, width=40, font=("Arial", 10))
        dk3_entry.grid(row=3, column=1, padx=5, pady=5)

        dk4_info = Label(frame, text="İlişki değerini giriniz:", font=("Arial", 10))
        dk4_info.grid(row=4, column=0, padx=5, pady=5)

        dk4_entry = Entry(frame, width=40, font=("Arial", 10))
        dk4_entry.grid(row=4, column=1, padx=5, pady=5)

        save_button = Button(frame, text="Kaydet", font=("Arial", 10),
                             command=lambda: insert_relation_value(dk2_entry.get(), dk3_entry.get(), dk4_entry.get(),
                                                                   dk1_entry.get()))
        save_button.grid(row=5, column=0, padx=5, pady=5)

        create_button = Button(frame, text="Tablo 1 oluştur", font=("Arial", 10),
                               command=lambda: create_table1())
        create_button.grid(row=5, column=1, padx=5, pady=5)

        frame5.pack(side=LEFT, fill=BOTH, expand=True, padx=5, pady=5)

    elif targetf == 6:
        frame6 = Frame(frame2)
        show_other_frames.active_frames.append(frame6)
        frame6_title = Label(frame6, text="Öğrenci Listesi", font=("Arial", 12))
        frame6_title.place(x=10, y=10)

        display_data_in_treeview("Students", frame6, 10, 40, selected_id)

        frame6.pack(side=LEFT, fill=BOTH, expand=True, padx=5, pady=5)

    elif targetf == 7:
        frame7 = Frame(frame2)
        show_other_frames.active_frames.append(frame7)
        frame7_title = Label(frame7, text="Değerlendirme Kriterleri ve Ağırlıkları Giriş Ekranı", font=("Arial", 12))
        frame7_title.place(x=0, y=10)

        kri_frame = Frame(frame7)
        kri_frame.place(x=0, y=50)
        lesson_id = selected_id
        frame = Frame(frame7)
        frame.place(x=100, y=400)

        def create_criteria_entries(num_criteria):
            # Giriş alanlarını temizle
            for widget in kri_frame.winfo_children():
                widget.destroy()

            entries = []

            Label(frame, text="Kriter Adı", font=("Arial", 10, "bold")).grid(row=0, column=1, padx=20, pady=5)
            Label(frame, text="Ağırlık (%)", font=("Arial", 10, "bold")).grid(row=0, column=2, padx=10, pady=5)

            for i in range(num_criteria):
                kriter_var = StringVar()
                agirlik_var = StringVar()

                Label(frame, text=f"Kriter {i + 1}:").grid(row=i + 1, column=0, sticky="e", padx=5, pady=5)
                kriter_entry = Entry(frame, textvariable=kriter_var, width=20)
                kriter_entry.grid(row=i + 1, column=1, padx=5, pady=5)
                agirlik_entry = Entry(frame, textvariable=agirlik_var, width=10)
                agirlik_entry.grid(row=i + 1, column=2, padx=5, pady=5)

                entries.append((lesson_id, kriter_var, agirlik_var))

            Button(frame, text="Kaydet", command=lambda: save_data1(entries)).grid(row=8, column=1, padx=5, pady=5)

        def set_criteria_count():
            try:
                num_criteria = int(criteria_count_entry.get())
                if num_criteria < 5:
                    raise ValueError("Minimum 5 kriter gereklidir.")
                create_criteria_entries(num_criteria)
            except ValueError as e:
                messagebox.showerror("Hata", f"Geçerli bir sayı girin! Hata: {str(e)}")

        def save_data1(criteria_data):

            total = sum(int(entry[2].get() or 0) for entry in criteria_data)
            if total != 100:
                messagebox.showerror("Hata", f"Toplam ağırlık 100 olmalıdır! Şu anki toplam: {total}")
                return

            conn = get_connection("RelationMatrix")
            conn.autocommit = True
            cursor = conn.cursor()

            data = [(int(entry[0]), entry[1].get(), int(entry[2].get()))
                    for entry in criteria_data if entry[1].get() and entry[2].get()]

            for lesson_id, criterion, weight in data:
                cursor.execute('''
                            INSERT INTO EvaluationCriteria (LessonID, Criteria, Weight)
                            VALUES (?, ?, ?);
                        ''', lesson_id, criterion, weight)

            create_table3()
            conn.close()

        dk_count = Label(kri_frame, text="Kaç Adet Değerlendirme Kriteri girilecek?.(Minimum 5 olmalı.)",
                         font=("Arial", 10))
        dk_count.grid(row=3, column=0, padx=5, pady=5)
        criteria_count_entry = Entry(kri_frame, width=5)
        criteria_count_entry.grid(row=4, column=0, padx=5, pady=5)

        Button(kri_frame, text="Kriterleri Oluştur", font=("Arial", 10),
               command=set_criteria_count).grid(row=5, column=0, padx=5, pady=5)

        frame7.pack(side=LEFT, fill=BOTH, expand=True, padx=5, pady=5)

    elif targetf == 8:
        frame8 = Frame(frame2)
        show_other_frames.active_frames.append(frame8)
        frame8_title = Label(frame8, text="Öğrenci Not Giriş Ekranı", font=("Arial", 12))
        frame8_title.grid(row=0, column=0, padx=5, pady=5)
        frame8.pack(side=LEFT, fill=BOTH, expand=True, padx=5, pady=5)

        ogr_no_label = Label(frame8, text="Öğrenci No giriniz:", font=("Arial", 10))
        ogr_no_label.grid(row=1, column=0, padx=5, pady=5)

        no_entry = Entry(frame8, width=40, font=("Arial", 10))
        no_entry.grid(row=1, column=1, padx=5, pady=5)

        conn = get_connection("RelationMatrix")
        conn.autocommit = True
        cursor = conn.cursor()

        query = """
                SELECT [Criteria]
                FROM [RelationMatrix].[dbo].[EvaluationCriteria]
                WHERE [LessonID] = ?
            """

        cursor.execute(query, (selected_id,))
        rows = cursor.fetchall()

        criteria_list = [row.Criteria for row in rows]
        entries = [no_entry.get(), selected_id]

        kriter_var_list = []

        for idx, criteria in enumerate(criteria_list, start=2):
            kriter_var = StringVar()
            Label(frame8, text=f"{criteria}:", font=("Arial", 10)).grid(row=idx, column=0, sticky="e", padx=5, pady=5)
            kriter_entry = Entry(frame8, textvariable=kriter_var, width=20, font=("Arial", 10))
            kriter_entry.grid(row=idx, column=1, padx=5, pady=5)
            kriter_var_list.append(kriter_var)

        Button(frame8, text="Kaydet", command=lambda: save_data(no_entry, kriter_var_list, criteria_list, frame8)).grid(
            row=len(criteria_list) + 2, column=1, padx=5, pady=5)

    elif targetf == 9:
        frame9 = Frame(frame2)
        show_other_frames.active_frames.append(frame9)
        frame5_title = Label(frame9, text="Ders Çıktısı ve Değerlendirme Kriteri Giriş Ekranı", font=("Arial", 12))
        frame5_title.place(x=0, y=10)

        kri_frame = Frame(frame9)
        kri_frame.place(x=0, y=40)

        frame = Frame(frame9)
        frame.place(x=100, y=400)

        show_button = Button(kri_frame, text="Tablo2 Görüntüle", font=("Arial", 10),
                             command=lambda: show_excel("table2", kri_frame, lesson_name))
        show_button.grid(row=0, column=0, padx=5, pady=5)

        dk1_info = Label(frame, text="Ders ID'sini giriniz:", font=("Arial", 10))
        dk1_info.grid(row=1, column=0, padx=5, pady=5)

        dk1_entry = Entry(frame, width=40, font=("Arial", 10))
        dk1_entry.grid(row=1, column=1, padx=5, pady=5)

        dk2_info = Label(frame, text="Ders Çıktısı\nID'sini giriniz:", font=("Arial", 10))
        dk2_info.grid(row=2, column=0, padx=5, pady=5)

        dk2_entry = Entry(frame, width=40, font=("Arial", 10))
        dk2_entry.grid(row=2, column=1, padx=5, pady=5)

        dk3_info = Label(frame, text="Değerlendirme\nKriteri giriniz:", font=("Arial", 10))
        dk3_info.grid(row=3, column=0, padx=5, pady=5)

        dk3_entry = Entry(frame, width=40, font=("Arial", 10))
        dk3_entry.grid(row=3, column=1, padx=5, pady=5)

        dk4_info = Label(frame, text="Değerlendirme Kriteri\n değeri giriniz:", font=("Arial", 10))
        dk4_info.grid(row=4, column=0, padx=5, pady=5)

        dk4_entry = Entry(frame, width=40, font=("Arial", 10))
        dk4_entry.grid(row=4, column=1, padx=5, pady=5)

        save_button = Button(frame, text="Kaydet", font=("Arial", 10),
                             command=lambda: insert_evaluation_relation_value(dk2_entry.get(), dk3_entry.get(),
                                                                              dk4_entry.get(), dk1_entry.get()))
        save_button.grid(row=5, column=0, padx=5, pady=5)

        create_button = Button(frame, text="Tablo 2 Ekle", font=("Arial", 10),
                               command=lambda: create_table2())
        create_button.grid(row=5, column=1, padx=5, pady=5)

        frame9.pack(side=LEFT, fill=BOTH, expand=True, padx=5, pady=5)

    def insert_evaluation_relation_value(course_outcome_id, criteria, relation_value, lesson_id):
        conn = get_connection("RelationMatrix")
        conn.autocommit = True
        cursor = conn.cursor()

        if not validate_input(relation_value):
            return

        cursor.execute('''
            INSERT INTO CourseEvaluationRelations (CourseOutcomeID, Criteria, RelationValue, LessonID)
            VALUES (?, ?, ?, ?);
        ''', course_outcome_id, criteria, relation_value, lesson_id)

        conn.close()

     

    def save_data(no_entry, kriter_var_list, criteria_list, frame8):
        # Extract student number and lesson ID
        student_no = no_entry.get()
        lesson_id = selected_id

        if not student_no.isdigit():
            messagebox.showerror("Hata", "Geçerli bir öğrenci numarası giriniz.")
            return

        criteria_values = [kriter_var.get() for kriter_var in kriter_var_list]

        if not all(value for value in criteria_values):
            messagebox.showerror("Hata", "Tüm kriterler için geçerli bir değer giriniz.")
            return

        conn = get_connection("RelationMatrix")
        cursor = conn.cursor()
        conn.autocommit = True

        # Prepare SQL query
        columns = ["Student", "lesson_id"] + [f"[{c}]" for c in criteria_list]
        columns_str = ', '.join(columns)
        placeholders = ', '.join(['?'] * len(columns))

        query = f"""
            INSERT INTO Students ({columns_str})
            VALUES ({placeholders})
        """

        try:
            # Prepare parameters for SQL query
            parameters = [student_no, lesson_id] + criteria_values

            cursor.execute(query, parameters)

            info_label = Label(frame8, text="Veriler başarıyla kaydedildi.", font=("Arial", 10))
            info_label.grid(row=len(criteria_list) + 3, column=1, padx=5, pady=5)


            for kriter_var in kriter_var_list:
                kriter_var.set("")
            no_entry.delete(0, END)

        except Exception as e:
            print("Hata:", e)
            info_label = Label(frame8, text="Bir hata oluştu. Veriler kaydedilemedi.", font=("Arial", 10), fg="red")
            info_label.grid(row=len(criteria_list) + 3, column=1, padx=5, pady=5)

        finally:
            conn.close()

check_database()
check_tables()
create_table4()
create_notes()
create_table5()
frame1.tkraise()
root.mainloop()