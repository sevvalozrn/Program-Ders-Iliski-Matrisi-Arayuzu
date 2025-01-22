import pyodbc
from openpyxl import Workbook
from openpyxl.comments import Comment

def get_connection(database=None):
    connection_string = (
        "Driver={ODBC Driver 17 for SQL Server};"
        "Server=ASUS\\SQLEXPRESS;"
        f"Database={database or 'master'};"
        "Trusted_Connection=yes;"
    )
    return pyodbc.connect(connection_string)

def check_database():
    conn = get_connection("master")
    conn.autocommit = True
    cursor = conn.cursor()

    cursor.execute("SELECT database_id FROM sys.databases WHERE Name = 'RelationMatrix'")
    database_exists = cursor.fetchone()

    if database_exists:
        print("There is already a database called RelationMatrix")
    else:
        cursor.execute("CREATE DATABASE RelationMatrix")
        print("RelationMatrix database created")

    conn.close()

def check_tables():
    conn = get_connection("RelationMatrix")
    conn.autocommit = True
    cursor = conn.cursor()

    cursor.execute("""
        IF NOT EXISTS (SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = 'Lessons')
        CREATE TABLE Lessons (
            id INT PRIMARY KEY IDENTITY(1,1),
            name NVARCHAR(255) NOT NULL
        );
    """)
    print("Lessons table checked/created.")

    tables = ["CourseOutcomes", "ProgramOutcomes", "ProgramCourseRelations", "EvaluationCriteria",
              "CourseEvaluationRelations"]
    for table in tables:
        cursor.execute(f"SELECT COUNT(*) FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = '{table}'")
        table_exists = cursor.fetchone()[0]
        if table_exists:
            print(f"{table} table already exists.")
        else:
            if table == 'CourseOutcomes':
                cursor.execute('''
                    CREATE TABLE CourseOutcomes (
                        id INT PRIMARY KEY IDENTITY(1,1),
                        data TEXT NOT NULL,
                        LessonID INT NOT NULL,
                        FOREIGN KEY (LessonID) REFERENCES Lessons(id)
                    );
                ''')
            elif table == 'ProgramOutcomes':
                cursor.execute('''
                    CREATE TABLE ProgramOutcomes (
                        id INT PRIMARY KEY IDENTITY(1,1),
                        data TEXT NOT NULL,
                        LessonID INT NOT NULL,
                        FOREIGN KEY (LessonID) REFERENCES Lessons(id)
                    );
                ''')
            elif table == 'ProgramCourseRelations':
                cursor.execute('''
                    CREATE TABLE ProgramCourseRelations (
                        ProgramOutcomeID INT NOT NULL,
                        CourseOutcomeID INT NOT NULL,
                        RelationValue FLOAT NOT NULL CHECK (RelationValue BETWEEN 0 AND 1),
                        LessonID INT NOT NULL,
                        PRIMARY KEY (ProgramOutcomeID, CourseOutcomeID, LessonID),
                        FOREIGN KEY (ProgramOutcomeID) REFERENCES ProgramOutcomes(id),
                        FOREIGN KEY (CourseOutcomeID) REFERENCES CourseOutcomes(id),
                        FOREIGN KEY (LessonID) REFERENCES Lessons(id)
                    );
                ''')
            elif table == 'EvaluationCriteria':
                cursor.execute('''
                    CREATE TABLE EvaluationCriteria (
                        Criteria NVARCHAR(25) NOT NULL,
                        Weight INT NOT NULL,
                        LessonID INT NOT NULL,
                        PRIMARY KEY (Criteria, LessonID),
                        FOREIGN KEY (LessonID) REFERENCES Lessons(id)
                    );
                ''')
            elif table == 'CourseEvaluationRelations':
                cursor.execute('''
                    CREATE TABLE CourseEvaluationRelations (
                        CourseOutcomeID INT NOT NULL,
                        Criteria NVARCHAR(25),
                        RelationValue INT NOT NULL,
                        LessonID INT NOT NULL,
                        PRIMARY KEY (CourseOutcomeID, Criteria, LessonID),
                        FOREIGN KEY (CourseOutcomeID) REFERENCES CourseOutcomes(id),
                        FOREIGN KEY (Criteria, LessonID) REFERENCES EvaluationCriteria(Criteria, LessonID),
                        FOREIGN KEY (LessonID) REFERENCES Lessons(id)
                    );
                ''')
            print(f"{table} table created.")

    conn.close()

def insert_data_into_table(table_name, data, lesson_id):
    conn = get_connection("RelationMatrix")
    conn.autocommit = True
    cursor = conn.cursor()

    if table_name == 'CourseOutcomes':
        cursor.execute("INSERT INTO CourseOutcomes (data, LessonID) VALUES (?, ?)", data, lesson_id)
    elif table_name == 'ProgramOutcomes':
        cursor.execute("INSERT INTO ProgramOutcomes (data, LessonID) VALUES (?, ?)", data, lesson_id)
    
    conn.close()


def fetch_relations():
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    query = """
    SELECT ProgramOutcomeID, CourseOutcomeID, RelationValue, LessonID
    FROM ProgramCourseRelations;
    """
    cursor.execute(query)

    relations = cursor.fetchall()

    cursor.close()
    conn.close()

    return relations


def fetch_evaluation_relations():
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    query = """
    SELECT CourseOutcomeID, Criteria, RelationValue, LessonID
    FROM CourseEvaluationRelations;
    """
    cursor.execute(query)

    relations = cursor.fetchall()

    cursor.close()
    conn.close()

    return relations


def fetch_table_data(table_name):
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    query = f"SELECT id, data, LessonID FROM {table_name};"
    cursor.execute(query)

    data = cursor.fetchall()

    cursor.close()
    conn.close()

    return data


def fetch_evaluation_data():
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    query = "SELECT Criteria, Weight, LessonID FROM EvaluationCriteria;"
    cursor.execute(query)
    data = cursor.fetchall()

    cursor.close()
    conn.close()

    return data


def fetch_student_data():
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    query = "SELECT Student, lesson_id FROM Students;"
    cursor.execute(query)
    students = cursor.fetchall()

    cursor.close()
    conn.close()

    return students


def fetch_success_rate():
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    query = "SELECT student_id, success_rate, lesson_id FROM Table4;"
    cursor.execute(query)
    success_rates = cursor.fetchall()

    cursor.close()
    conn.close()

    return success_rates

def fetch_lesson_names():
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    query = "SELECT id, name FROM Lessons;"
    cursor.execute(query)

    lesson_names = dict(cursor.fetchall())  

    cursor.close()
    conn.close()

    return lesson_names

def get_input_and_insert_relations(lesson_id):
    if not lesson_id:
        print("Lesson ID is required to insert relations.")
        return

    print("Enter 'q' to quit.")
    while True:
        program_outcome_id = input("Enter Program Outcome ID: ").strip()
        if program_outcome_id.lower() == 'q':
            print("Exiting the program.")
            break

        course_outcome_id = input("Enter Course Outcome ID: ").strip()
        if course_outcome_id.lower() == 'q':
            print("Exiting the program.")
            break

        relation_value = input("Enter Relation Value (0-1): ").strip()
        if relation_value.lower() == 'q':
            print("Exiting the program.")
            break

        try:
            relation_value = float(relation_value)
            if 0 <= relation_value <= 1:
                insert_relation_value(program_outcome_id, course_outcome_id, relation_value, lesson_id)
                print(f"Relation inserted for Lesson ID {lesson_id}.")
            else:
                print("Please enter a relation value between 0 and 1.")
        except ValueError:
            print("Invalid input for relation value. Please enter a valid number between 0 and 1.")


def get_input_and_insert_evaluation_relations(lesson_id):
    if not lesson_id:
        print("Lesson ID is required to insert evaluation relations.")
        return

    print("Enter 'q' to quit.")
    while True:
        course_outcome_id = input("Enter Course Outcome ID: ").strip()
        if course_outcome_id.lower() == 'q':
            print("Exiting the program.")
            break

        evaluation_criteria = input("Enter Evaluation Criteria: ").strip()
        if evaluation_criteria.lower() == 'q':
            print("Exiting the program.")
            break

        relation_value = input("Enter Relation Value (0/1): ").strip()
        if relation_value.lower() == 'q':
            print("Exiting the program.")
            break

        try:
            relation_value = int(relation_value)
            if relation_value in [0, 1]:
                insert_evaluation_relation_value(course_outcome_id, evaluation_criteria, relation_value, lesson_id)
                print(f"Relation inserted for Lesson ID {lesson_id}.")
            else:
                print("Please enter a relation value of 0 or 1.")
        except ValueError:
            print("Invalid input for relation value. Please enter 0 or 1.")


def insert_relation_value(program_outcome_id, course_outcome_id, relation_value, lesson_id):
    conn = get_connection("RelationMatrix")
    conn.autocommit = True
    cursor = conn.cursor()

    cursor.execute('''
        INSERT INTO ProgramCourseRelations (ProgramOutcomeID, CourseOutcomeID, RelationValue, LessonID)
        VALUES (?, ?, ?, ?);
    ''', program_outcome_id, course_outcome_id, relation_value, lesson_id)

    print(f"Relation between ProgramOutcome {program_outcome_id} and CourseOutcome {course_outcome_id} for Lesson {lesson_id} has been inserted.")
    conn.close()


def insert_evaluation_relation_value(course_outcome_id, criteria, relation_value, lesson_id):
    conn = get_connection("RelationMatrix")
    conn.autocommit = True
    cursor = conn.cursor()

    cursor.execute('''
        INSERT INTO CourseEvaluationRelations (CourseOutcomeID, Criteria, RelationValue, LessonID)
        VALUES (?, ?, ?, ?);
    ''', course_outcome_id, criteria, relation_value, lesson_id)

    print(f"Relation between CourseOutcome {course_outcome_id} and EvaluationCriteria {criteria} for Lesson {lesson_id} has been inserted.")

    conn.close()


def clear_relations():
    conn = get_connection("RelationMatrix")
    conn.autocommit = True
    cursor = conn.cursor()

    cursor.execute('DELETE FROM CourseEvaluationRelations;')
    print("CourseEvaluationRelations table cleared.")

    cursor.execute('DELETE FROM EvaluationCriteria;')
    print("EvaluationCriteria table cleared.")

    conn.close()


def get_input_and_insert_table(table_name, lesson_id):
    print("Enter 'q' to quit.")
    while True:
        data = input(f"Enter data for {table_name}: ").strip()
        if data == 'q':
            print(f"Exiting {table_name} input.")
            break

        insert_data_into_table(table_name, data, lesson_id)
        print(f"Data has been inserted into {table_name}.")


def get_evaluation_criteria_and_insert(lesson_id):
    print("Enter evaluation criteria and their weights. The total weight must be 100.")
    criteria_data = []
    total_weight = 0

    conn = get_connection("RelationMatrix")
    conn.autocommit = True
    cursor = conn.cursor()

    while True:
        criterion = input("Enter Criterion (or 'q' to quit): ").strip()
        if criterion.lower() == 'q':
            break

        weight = input(f"Enter Weight for {criterion}: ").strip()
        try:
            weight = int(weight)
            if weight < 0:
                print("Weight must be a positive integer.")
                continue
        except ValueError:
            print("Invalid weight. Please enter a positive integer.")
            continue

        total_weight += weight
        if total_weight > 100:
            print(f"Total weight exceeded 100 (current total: {total_weight}). Adjust your inputs.")
            total_weight -= weight
            continue

        criteria_data.append((criterion, weight))

        if total_weight == 100:
            break

        print(f"Current total weight: {total_weight}. You need {100 - total_weight} more.")

    if total_weight < 100:
        print(f"Total weight is {total_weight}, which is less than 100. Please try again.")
        return

    conn = get_connection("RelationMatrix")
    conn.autocommit = True
    cursor = conn.cursor()

    for criterion, weight in criteria_data:
        cursor.execute('''
            INSERT INTO EvaluationCriteria (Criteria, Weight, LessonID)
            VALUES (?, ?, ?);
        ''', criterion, weight, lesson_id)  

    print("Evaluation criteria have been successfully inserted into the database.")
    conn.close()


is_table_created = False
def create_table1():
    workbook = Workbook()
    lesson_names = fetch_lesson_names()  # LessonID ve Lesson name bilgileri

    program_outcomes = fetch_table_data("ProgramOutcomes")
    course_outcomes = fetch_table_data("CourseOutcomes")

    program_row_count = len(program_outcomes)
    course_row_count = len(course_outcomes)

    # LessonID'lere göre her derse sheet oluşturma  
    for lesson_id, lesson_name in lesson_names.items():
        sheet = workbook.create_sheet(lesson_name) 

        sheet.merge_cells('A1:B1')
        sheet['A1'] = f"Table 1 - {lesson_name}"
        sheet.merge_cells('A2:B2')
        sheet['A2'] = "Program Outcomes"
        sheet.merge_cells(start_row=1, start_column=3, end_row=1, end_column=course_row_count + 2)
        sheet['C1'] = "Course Outcomes"

        for i, (program_id, program_text, program_lesson_id) in enumerate(program_outcomes, start=1):
            if program_lesson_id == lesson_id:  # LessonID'ye göre filtreleme işlemi
                sheet.merge_cells(f"A{i + 2}:B{i + 2}")
                cell = sheet[f"A{i + 2}"]
                cell.value = program_id
                comment = Comment(program_text, "Database")
                cell.comment = comment

        for j, (course_id, course_text, course_lesson_id) in enumerate(course_outcomes, start=1):
            if course_lesson_id == lesson_id:  # LessonID'ye göre filtreleme işlemi
                c = sheet.cell(row=2, column=j + 2)
                c.value = course_id
                comment = Comment(course_text, "Database")
                c.comment = comment

        # Program-Course relations'ın sheet'e eklenmesi
        relations = fetch_relations()
        for relation in relations:
            program_outcome_id, course_outcome_id, relation_value, relation_lesson_id = relation
            if relation_lesson_id == lesson_id:  
                row = program_outcome_id + 2
                col = course_outcome_id + 2
                sheet.cell(row=row, column=col, value=relation_value)

        # Course outcomes'a ilişkin her program outcome için toplam rel value hesaplama
        for i in range(3, program_row_count + 3):
            course_outcomes_count = sum(
                1 for (course_id, _, course_lesson_id) in course_outcomes if course_lesson_id == lesson_id
            )

            total = sum(sheet.cell(row=i, column=j).value or 0 for j in range(3, course_row_count + 3))
            result = round((total / course_outcomes_count), 2) if course_outcomes_count > 0 else 0
            sheet.cell(row=i, column=course_row_count + 3, value=result)

        sheet.cell(row=2, column=course_row_count + 3, value="Rel Value")
    del workbook['Sheet']
    
    workbook.save(filename="table1.xlsx") 


def create_table2():
    workbook = Workbook()
    lesson_names = fetch_lesson_names()  

    for lesson_id, lesson_name in lesson_names.items():
        sheet = workbook.create_sheet(lesson_name)

        # Başlıkların eklenmesi
        sheet.merge_cells('A1:B1')
        sheet['A1'] = f"Table 2 - {lesson_name}" 
        sheet.merge_cells('A2:B2')
        sheet['A2'] = "Course Outcomes"

        # İlgili dersin Course Outcomes verilerinin getirilmesi 
        course_outcomes = fetch_table_data("CourseOutcomes")
        filtered_course_outcomes = [
            (course_id, course_text) for course_id, course_text, course_lesson_id in course_outcomes
            if course_lesson_id == lesson_id
        ]
        for i, (course_id, course_text) in enumerate(filtered_course_outcomes, start=1):
            sheet.merge_cells(f"A{i + 2}:B{i + 2}")
            cell = sheet[f"A{i + 2}"]
            cell.value = course_id
            comment = Comment(course_text, "Database")
            cell.comment = comment

        evaluation_criteria = fetch_evaluation_data()
        filtered_criteria = [
            (criteria, weight) for criteria, weight, criteria_lesson_id in evaluation_criteria
            if criteria_lesson_id == lesson_id
        ]
        for index, (criteria, weight) in enumerate(filtered_criteria, start=3):
            sheet.cell(row=1, column=index, value=weight)  # Ağırlık
            sheet.cell(row=2, column=index, value=criteria)  # Kriter adı

        evaluation_relations = fetch_evaluation_relations()
        for relation in evaluation_relations:
            course_outcome_id, criteria, relation_value, relation_lesson_id = relation
            if relation_lesson_id == lesson_id:
                row = None
                for i, (course_id, _) in enumerate(filtered_course_outcomes, start=3):
                    if course_outcome_id == course_id:
                        row = i
                        break

                col = None
                for j, (crit, _) in enumerate(filtered_criteria, start=3):
                    if crit == criteria:
                        col = j
                        break

                if row and col:
                    sheet.cell(row=row, column=col, value=relation_value)

        # Her satır için toplam hesaplama
        total_col = len(filtered_criteria) + 3
        sheet.cell(row=2, column=total_col, value="Total")
        for row_idx in range(3, len(filtered_course_outcomes) + 3):
            total = sum(sheet.cell(row=row_idx, column=col_idx).value or 0 for col_idx in range(3, total_col))
            sheet.cell(row=row_idx, column=total_col, value=total)

    del workbook['Sheet']
    workbook.save(filename="table2.xlsx")


def create_table3():
    workbook = Workbook()

    # Ders ID'lerine göre veri filtreleme işlemi
    lesson_names = fetch_lesson_names() 
    course_evaluation_relations = fetch_evaluation_relations() 
    course_outcomes = fetch_table_data("CourseOutcomes") 
    evaluation_criteria = fetch_evaluation_data()  

    # EvaluationCriteria verilerinin LessonID bazında gruplanması
    criteria_weights = {
        (lesson_id, criteria): weight
        for criteria, weight, lesson_id in evaluation_criteria
    }

    for lesson_id, lesson_name in lesson_names.items():
        sheet = workbook.create_sheet(lesson_name)

        sheet.merge_cells('A1:B1')
        sheet['A1'] = f"Table 3 - {lesson_name}"
        sheet['C1'] = "Weighted Evaluation"

        sheet.merge_cells('A2:B2')
        sheet['A2'] = "Course Outcomes"

        filtered_criteria = [
            criteria for _, criteria, _, relation_lesson_id in course_evaluation_relations
            if relation_lesson_id == lesson_id and (relation_lesson_id, criteria) in criteria_weights
        ]

        for col, criteria in enumerate(set(filtered_criteria), start=3):
            sheet.cell(row=2, column=col, value=criteria)

        total_col = len(set(filtered_criteria)) + 3
        sheet.cell(row=2, column=total_col, value="Total")

        weighted_data = {}
        filtered_course_evaluation_relations = [
            (course_outcome_id, criteria, relation_value)
            for course_outcome_id, criteria, relation_value, relation_lesson_id in course_evaluation_relations
            if relation_lesson_id == lesson_id
        ]

        filtered_course_outcomes = [
            (course_outcome_id, program_text)
            for course_outcome_id, program_text, course_lesson_id in course_outcomes
            if course_lesson_id == lesson_id
        ]

        for course_outcome_id, criteria, relation_value in filtered_course_evaluation_relations:
            weight = criteria_weights.get((lesson_id, criteria), 0)  
            weighted_value = (relation_value * weight) / 100

            if course_outcome_id not in weighted_data:
                weighted_data[course_outcome_id] = {}

            weighted_data[course_outcome_id][criteria] = weighted_value

        for row_idx, (course_outcome_id, program_text) in enumerate(filtered_course_outcomes, start=3):
            sheet.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
            cell = sheet.cell(row=row_idx, column=1, value=course_outcome_id)
            comment = Comment(program_text, "Database")
            cell.comment = comment

            total = 0
            # Her kriterin değerinin eklenmesi
            for col_idx, criteria in enumerate(set(filtered_criteria), start=3):
                value = weighted_data.get(course_outcome_id, {}).get(criteria, 0)
                sheet.cell(row=row_idx, column=col_idx, value=value)
                total += value

            sheet.cell(row=row_idx, column=total_col, value=total)

    if 'Sheet' in workbook.sheetnames:
        del workbook['Sheet']

    workbook.save("table3.xlsx")


def save_table3_to_database():
    # Veri çekme işlemleri
    lesson_names = fetch_lesson_names() 
    course_evaluation_relations = fetch_evaluation_relations()
    course_outcomes = fetch_table_data("CourseOutcomes")
    evaluation_criteria = fetch_evaluation_data()

    if evaluation_criteria and isinstance(evaluation_criteria[0], tuple):
        criteria_weights = {criteria: weight for criteria, weight in evaluation_criteria}
    else:
        criteria_weights = {item[0]: item[1] for item in evaluation_criteria if len(item) >= 2}

    conn = get_connection("RelationMatrix") 
    cursor = conn.cursor()

    cursor.execute("SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = 'Table3';")
    existing_table = cursor.fetchone()
    if existing_table:
        cursor.execute("DROP TABLE Table3;")

    criteria_columns = ", ".join(f'"{criteria}" FLOAT' for criteria in criteria_weights.keys())
    cursor.execute(f"""
        CREATE TABLE Table3 (
            id INT IDENTITY(1,1) PRIMARY KEY,
            lesson_id INT NOT NULL,
            course_outcome_id INT NOT NULL,
            total_score FLOAT NOT NULL,
            {criteria_columns}
        );
    """)

    # Verilerin lesson_id'ye göre işlenmesi
    for lesson_id, lesson_name in lesson_names.items():
        filtered_course_evaluation_relations = [
            (course_outcome_id, criteria, relation_value)
            for course_outcome_id, criteria, relation_value, relation_lesson_id in course_evaluation_relations
            if relation_lesson_id == lesson_id
        ]

        filtered_course_outcomes = [
            (course_outcome_id, program_text)
            for course_outcome_id, program_text, course_lesson_id in course_outcomes
            if course_lesson_id == lesson_id
        ]

        weighted_data = {}
        for course_outcome_id, criteria, relation_value in filtered_course_evaluation_relations:
            weight = criteria_weights.get(criteria, 0)
            weighted_value = (relation_value * weight) / 100
            if course_outcome_id not in weighted_data:
                weighted_data[course_outcome_id] = {}
            weighted_data[course_outcome_id][criteria] = weighted_value

        # Veritabanına ekleme işlemi
        for course_outcome_id, program_text in filtered_course_outcomes:
            total_score = 0
            criteria_values = []
            for criteria in criteria_weights.keys():
                value = weighted_data.get(course_outcome_id, {}).get(criteria, 0)
                criteria_values.append(value)
                total_score += value

            columns = ", ".join(f'"{criteria}"' for criteria in criteria_weights.keys())
            placeholders = ", ".join("?" for _ in criteria_values)

            cursor.execute(f"""
                INSERT INTO Table3 (lesson_id, course_outcome_id, total_score, {columns})
                VALUES (?, ?, ?, {placeholders})
            """, (lesson_id, course_outcome_id, total_score, *criteria_values))

    conn.commit()
    conn.close()


def create_notes():
    workbook = Workbook()
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    # Ders adlarını almak için fetch_lesson_names fonksiyonu kullanımı
    lesson_names = fetch_lesson_names()

    # Öğrenciler tablosundaki derslerin alınması
    cursor.execute("SELECT DISTINCT lesson_id FROM Students;")
    lesson_ids = [row[0] for row in cursor.fetchall()]

    if not lesson_ids:
        print("No lessons found in Students table.")
        return

    for lesson_id in lesson_ids:
        lesson_name = lesson_names.get(lesson_id, "Unknown Lesson")

        sheet = workbook.create_sheet(title=f"Lesson {lesson_id}")
        sheet.merge_cells('A1:B1')
        sheet['A1'] = f"Table Note - {lesson_name}"  
        sheet['C1'] = "Notes"

        cursor.execute("""
            SELECT ec.Criteria, ec.Weight
            FROM EvaluationCriteria ec
            WHERE ec.LessonID = ?;
        """, (lesson_id,))
        criteria_weights = {row[0]: row[1] for row in cursor.fetchall()}

        if not criteria_weights:
            print(f"No criteria found for Lesson ID {lesson_id}.")
            continue

        criteria_columns = [f"[{criterion}]" for criterion in criteria_weights.keys()]
        columns_str = ", ".join(criteria_columns)

        cursor.execute(f"""
            SELECT s.Student, s.lesson_id, {columns_str}
            FROM Students s
            WHERE s.lesson_id = ?;
        """, (lesson_id,))
        rows = cursor.fetchall()

        if not rows:
            print(f"No data found for Lesson ID {lesson_id}.")
            continue

        columns = ["Student "] + list(criteria_weights.keys()) + ["Average"]
        for col_idx, column_name in enumerate(columns, start=1):
            sheet.cell(row=2, column=col_idx, value=column_name)

        for row_idx, row in enumerate(rows, start=3):
            student_id = row[0]
            sheet.cell(row=row_idx, column=1, value=student_id)

            total_score = 0
            weight_sum = 0
            for col_idx, (criterion, value) in enumerate(zip(criteria_weights.keys(), row[2:]), start=2):
                if value is not None:
                    sheet.cell(row=row_idx, column=col_idx, value=value)
                    weight = criteria_weights.get(criterion, 0)
                    total_score += value * weight
                    weight_sum += weight
                else:
                    sheet.cell(row=row_idx, column=col_idx, value=0)

            average = total_score / 100 if weight_sum == 100 else total_score / weight_sum
            sheet.cell(row=row_idx, column=len(columns), value=round(average, 2))

    del workbook['Sheet']
    workbook.save("notlar.xlsx")
    conn.close()


def create_students_table(lesson_id):
    global is_table_created

    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    # Kullanıcının seçtiği lesson_id'ye ait kriterlerin alınması
    criteria_query = """
    SELECT DISTINCT ec.Criteria 
    FROM EvaluationCriteria ec
    WHERE ec.LessonID = ?;
    """
    cursor.execute(criteria_query, (lesson_id,))
    criteria = [row[0] for row in cursor.fetchall()]

    if not criteria:
        print(f"No evaluation criteria found for Lesson ID {lesson_id}.")
        conn.close()
        return

    cursor.execute("SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = 'Students';")
    existing_table = cursor.fetchone()

    if existing_table:
        print("Students table already exists.")
    else:
        columns = ["Student INT", "lesson_id INT", "PRIMARY KEY (Student, lesson_id)"]
        for criterion in criteria:
            columns.append(f'[{criterion}] FLOAT')

        create_table_query = f"CREATE TABLE Students ({', '.join(columns)});"
        cursor.execute(create_table_query)
        conn.commit()
    
    for criterion in criteria:
        column_check_query = f"SELECT COLUMN_NAME FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = 'Students' AND COLUMN_NAME = '{criterion}'"
        cursor.execute(column_check_query)
        column_exists = cursor.fetchone()

        if not column_exists:
            alter_query = f"ALTER TABLE Students ADD [{criterion}] FLOAT"
            cursor.execute(alter_query)
            conn.commit()

    conn.close()
    is_table_created = True


def create_table4():
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    # Ders kriter ağırlıklarının çekilmesi
    cursor.execute("SELECT Criteria, Weight, LessonID FROM EvaluationCriteria;")
    evaluation_weights = {}
    for criteria, weight, lesson_id in cursor.fetchall():
        if lesson_id not in evaluation_weights:
            evaluation_weights[lesson_id] = {}
        evaluation_weights[lesson_id][criteria] = weight

    # Öğrenci verilerinin çekilmesi
    cursor.execute("SELECT * FROM Students;")
    student_columns = [column[0] for column in cursor.description]
    student_data = cursor.fetchall()

    # Ders çıktıları (CourseOutcomes) bilgilerinin çekilmesi
    cursor.execute("SELECT id, data, LessonID FROM CourseOutcomes;")
    course_outcomes = {}
    outcome_descriptions = {} 
    for outcome_id, data, lesson_id in cursor.fetchall():
        if lesson_id not in course_outcomes:
            course_outcomes[lesson_id] = []
        course_outcomes[lesson_id].append((outcome_id, data))
        outcome_descriptions[outcome_id] = data

    # Ders değerlendirme ilişkilerinin çekilmesi
    cursor.execute("SELECT CourseOutcomeID, Criteria, RelationValue, LessonID FROM CourseEvaluationRelations;")
    course_evaluation_relations = {}
    for outcome_id, criteria, relation_value, lesson_id in cursor.fetchall():
        if lesson_id not in course_evaluation_relations:
            course_evaluation_relations[lesson_id] = {}
        if outcome_id not in course_evaluation_relations[lesson_id]:
            course_evaluation_relations[lesson_id][outcome_id] = {}
        course_evaluation_relations[lesson_id][outcome_id][criteria] = relation_value

    cursor.execute("SELECT lesson_id, course_outcome_id, total_score FROM Table3;")
    total_scores = {}
    for lesson_id, course_outcome_id, total_score in cursor.fetchall():
        if lesson_id not in total_scores:
            total_scores[lesson_id] = {}
        total_scores[lesson_id][course_outcome_id] = total_score

    conn.close()

    lesson_names = fetch_lesson_names()

    workbook = Workbook()
    sheet_created = {}

    for student in student_data:
        student_lesson_id = student[student_columns.index("lesson_id")]
        student_id = student[student_columns.index("Student")] 
        lesson_name = lesson_names.get(student_lesson_id, f"Lesson {student_lesson_id}") 

        if student_lesson_id not in sheet_created:
            sheet = workbook.create_sheet(title=lesson_name)
            sheet_created[student_lesson_id] = sheet

            sheet.merge_cells("A1:B1")
            sheet["A1"] = f"{lesson_name} - Table 4"

            sheet["A2"] = "Student ID" 
            sheet["B2"] = "Course Outcomes"

            # Kriterler, toplam, maksimum başarı oranı başlıklarının oluşturulması
            headers = [*evaluation_weights[student_lesson_id].keys(), "Total", "Max", "% Success"]
            for col_index, header in enumerate(headers, start=3):
                sheet.cell(row=2, column=col_index, value=header)

        sheet = sheet_created[student_lesson_id]
        current_row = sheet.max_row + 1

        for outcome_id, outcome_text in course_outcomes[student_lesson_id]:
            sheet.cell(row=current_row, column=1, value=student_id)  
            sheet.cell(row=current_row, column=2, value=outcome_text) 

            sheet.cell(row=current_row, column=2).comment = Comment(outcome_descriptions[outcome_id], "System")

            row = []
            total = 0

            for criteria, weight in evaluation_weights[student_lesson_id].items():
                score = student[student_columns.index(criteria)]
                relation_value = course_evaluation_relations[student_lesson_id].get(outcome_id, {}).get(criteria, 0)

                if score is not None and relation_value != 0:

                    weighted_score = score * weight * relation_value / 100
                else:
                    weighted_score = 0  
                row.append(weighted_score)
                total += weighted_score

            max_score = total_scores.get(student_lesson_id, {}).get(outcome_id, 0) * 100
            success_rate = (total / max_score * 100) if max_score > 0 else 0

            row.extend([total, max_score, round(success_rate, 1)])

            for col_index, value in enumerate(row, start=3):
                sheet.cell(row=current_row, column=col_index, value=value)

            current_row += 1

    del workbook['Sheet']
    workbook.save("table4.xlsx")


def save_table4_to_database():
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    cursor.execute("SELECT Criteria, Weight, LessonID FROM EvaluationCriteria;")
    evaluation_weights = {}
    for criteria, weight, lesson_id in cursor.fetchall():
        if lesson_id not in evaluation_weights:
            evaluation_weights[lesson_id] = {}
        evaluation_weights[lesson_id][criteria] = weight

    cursor.execute("SELECT * FROM Students;")
    student_columns = [column[0] for column in cursor.description]
    student_data = cursor.fetchall()

    cursor.execute("SELECT id, data, LessonID FROM CourseOutcomes;")
    course_outcomes = {}
    for outcome_id, data, lesson_id in cursor.fetchall():
        if lesson_id not in course_outcomes:
            course_outcomes[lesson_id] = []
        course_outcomes[lesson_id].append(outcome_id)

    cursor.execute("SELECT CourseOutcomeID, Criteria, RelationValue, LessonID FROM CourseEvaluationRelations;")
    course_evaluation_relations = {}
    for outcome_id, criteria, relation_value, lesson_id in cursor.fetchall():
        if lesson_id not in course_evaluation_relations:
            course_evaluation_relations[lesson_id] = {}
        if outcome_id not in course_evaluation_relations[lesson_id]:
            course_evaluation_relations[lesson_id][outcome_id] = {}
        course_evaluation_relations[lesson_id][outcome_id][criteria] = relation_value

    cursor.execute("SELECT lesson_id, course_outcome_id, total_score FROM Table3;")
    total_scores = {}
    for lesson_id, course_outcome_id, total_score in cursor.fetchall():
        if lesson_id not in total_scores:
            total_scores[lesson_id] = {}
        total_scores[lesson_id][course_outcome_id] = total_score

    cursor.execute("SELECT * FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = 'Table4';")
    existing_table = cursor.fetchone()
    if existing_table:
        cursor.execute("DROP TABLE Table4;")

    columns_by_lesson = {lesson_id: ", ".join(f'"{criteria}" FLOAT' for criteria in evaluation_weights[lesson_id])
                         for lesson_id in evaluation_weights}

    cursor.execute(f"""
        CREATE TABLE Table4 (
            id INT IDENTITY(1,1) PRIMARY KEY,
            student_id INT NOT NULL,
            lesson_id INT NOT NULL,
            course_outcome_id INT NOT NULL,
            total_score FLOAT NOT NULL,
            max_score FLOAT,
            success_rate FLOAT
        );
    """)

    for student in student_data:
        student_lesson_id = student[student_columns.index("lesson_id")]
        student_id = student[student_columns.index("Student")]

        if student_lesson_id not in course_outcomes:
            continue 

        for outcome_id in course_outcomes[student_lesson_id]:
            total_score = 0
            criteria_values = []

            for criteria, weight in evaluation_weights[student_lesson_id].items():
                score = student[student_columns.index(criteria)]
                relation_value = course_evaluation_relations[student_lesson_id].get(outcome_id, {}).get(criteria, 0)

                if score is not None and relation_value != 0:
                    weighted_score = score * weight * relation_value / 100
                else:
                    weighted_score = 0

                criteria_values.append(weighted_score)
                total_score += weighted_score

            max_score = total_scores.get(student_lesson_id, {}).get(outcome_id, 0) * 100
            success_rate = (total_score / max_score * 100) if max_score > 0 else 0

            cursor.execute(f"""
                INSERT INTO Table4 (student_id, lesson_id, course_outcome_id, total_score, max_score, success_rate)
                VALUES (?, ?, ?, ?, ?, ?);
            """, (student_id, student_lesson_id, outcome_id, total_score, max_score, round(success_rate, 1)))

    conn.commit()
    conn.close()


def fetch_student_lessons(student_id):
    conn = get_connection("RelationMatrix")  
    cursor = conn.cursor()

    cursor.execute('''
        SELECT lesson_id
        FROM table4
        WHERE student_id = ?
    ''', (student_id,))

    lessons = cursor.fetchall()
    conn.close()

    return [lesson[0] for lesson in lessons]

def create_table5():
    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    cursor.execute("SELECT id, name FROM Lessons;")
    lessons = {row[0]: row[1] for row in cursor.fetchall()}

    cursor.execute("SELECT * FROM Students;")
    student_columns = [column[0] for column in cursor.description]
    student_data = cursor.fetchall()

    cursor.execute("SELECT id, data, LessonID FROM ProgramOutcomes;")
    program_outcomes = {}
    for program_id, data, lesson_id in cursor.fetchall():
        if lesson_id not in program_outcomes:
            program_outcomes[lesson_id] = {}
        program_outcomes[lesson_id][program_id] = data

    cursor.execute("SELECT id, data, LessonID FROM CourseOutcomes;")
    course_outcomes = {}
    for outcome_id, data, lesson_id in cursor.fetchall():
        if lesson_id not in course_outcomes:
            course_outcomes[lesson_id] = []
        course_outcomes[lesson_id].append((outcome_id, data))

    cursor.execute("SELECT ProgramOutcomeID, CourseOutcomeID, RelationValue, LessonID FROM ProgramCourseRelations;")
    relations = {}
    for program_id, course_id, relation_value, lesson_id in cursor.fetchall():
        if lesson_id not in relations:
            relations[lesson_id] = {}
        if program_id not in relations[lesson_id]:
            relations[lesson_id][program_id] = {}
        relations[lesson_id][program_id][course_id] = relation_value

    cursor.execute("SELECT student_id, course_outcome_id, success_rate FROM Table4;")
    success_rates = {}
    for student_id, course_outcome_id, success_rate in cursor.fetchall():
        if student_id not in success_rates:
            success_rates[student_id] = {}
        success_rates[student_id][course_outcome_id] = success_rate

    conn.close()
    workbook = Workbook()

    for lesson_id, lesson_name in lessons.items():
        sheet = workbook.create_sheet(title=f"{lesson_name}")
        sheet.merge_cells("A1:C1")
        sheet["A1"] = f"{lesson_name} - Table 5"

        sheet["A2"] = "Student ID"
        sheet["B2"] = "Program Outcomes"
        lesson_course_outcomes = course_outcomes.get(lesson_id, [])
        for idx, (_, course_outcome_text) in enumerate(lesson_course_outcomes, start=3):
            sheet.cell(row=2, column=idx, value=course_outcome_text)

        sheet.cell(row=2, column=len(lesson_course_outcomes) + 3, value="Success Rate")

        current_row = 3
        for student in student_data:
            student_id = student[student_columns.index("Student")]
            student_lesson_id = student[student_columns.index("lesson_id")]

            if student_lesson_id != lesson_id:
                continue 

            lesson_program_outcomes = program_outcomes.get(lesson_id, {})

            for program_outcome_id, program_outcome_text in lesson_program_outcomes.items():
                relation_values = []
                row_values = []

                for course_outcome_id, _ in lesson_course_outcomes:
                    relation_value = relations.get(lesson_id, {}).get(program_outcome_id, {}).get(course_outcome_id, 0)
                    success_rate = success_rates.get(student_id, {}).get(course_outcome_id, 0)
                    weighted_value = relation_value * success_rate
                    relation_values.append(relation_value)
                    row_values.append(weighted_value)

                total_success = sum(row_values)
                num_course_outcomes = len(lesson_course_outcomes)
                avg_success = total_success / num_course_outcomes if num_course_outcomes else 0
                avg_relation_value = sum(relation_values) / len(relation_values) if relation_values else 0
                ratio = avg_success / avg_relation_value if avg_relation_value else 0

                sheet.cell(row=current_row, column=1, value=student_id)  
                sheet.cell(row=current_row, column=2, value=program_outcome_text)
                for col_index, value in enumerate(row_values, start=3):
                    sheet.cell(row=current_row, column=col_index, value=round(value, 1))
                sheet.cell(row=current_row, column=len(lesson_course_outcomes) + 3, value=round(ratio, 1))

                current_row += 1

    del workbook["Sheet"]
    workbook.save("table5.xlsx")


def evaluation_criteria_and_insert_table5():
    print("Enter evaluation criteria and their weights. The total weight must be 100 for each lesson.")

    while True:
        lesson_id = input("Enter Lesson ID (or 'q' to quit): ").strip()
        if lesson_id.lower() == 'q':
            print("Exiting the program.")
            break

        criteria_data = []
        total_weight = 0

        while True:
            criterion = input("Enter Criterion (or 'q' to quit current lesson): ").strip()
            if criterion.lower() == 'q':
                print(f"Exiting input for Lesson {lesson_id}.")
                break

            weight = input(f"Enter Weight for {criterion}: ").strip()
            try:
                weight = int(weight)
                if weight < 0:
                    print("Weight must be a positive integer.")
                    continue
            except ValueError:
                print("Invalid weight. Please enter a positive integer.")
                continue

            total_weight += weight
            if total_weight > 100:
                print(f"Total weight exceeded 100 (current total: {total_weight}). Adjust your inputs.")
                total_weight -= weight
                continue

            criteria_data.append((lesson_id, criterion, weight))

            if total_weight == 100:
                print(f"Total weight for Lesson {lesson_id} reached 100. Moving to database insertion.")
                break

            print(f"Current total weight: {total_weight}. You need {100 - total_weight} more.")

        if total_weight < 100:
            print(f"Total weight for Lesson {lesson_id} is {total_weight}, which is less than 100. Please try again.")
            continue

        conn = get_connection("RelationMatrix")
        conn.autocommit = True
        cursor = conn.cursor()

        for lesson_id, criterion, weight in criteria_data:
            cursor.execute('''
                INSERT INTO EvaluationCriteria (LessonID, Criteria, Weight)
                VALUES (?, ?, ?);
            ''', lesson_id, criterion, weight)

        conn.close()
        print(f"Evaluation criteria for Lesson {lesson_id} have been successfully inserted into the database.")


is_table_created = True
def add_student(lesson_id):
    if not lesson_id:
        print("Lesson ID is required to add a student.")
        return

    conn = get_connection("RelationMatrix")
    cursor = conn.cursor()

    # Seçilen lesson_id'ye ait tüm kriterlerin alınması
    criteria_query = """
    SELECT DISTINCT ec.Criteria 
    FROM EvaluationCriteria ec
    WHERE ec.LessonID = ?;
    """
    cursor.execute(criteria_query, (lesson_id,))
    criteria = [row[0] for row in cursor.fetchall()]

    if not criteria:
        print(f"No evaluation criteria found for Lesson ID {lesson_id}. Please add criteria first.")
        conn.close()
        return

    criteria = list(set(criteria))

    while True:
        student_number = input("Enter Student Number (or 'q' to quit): ").strip()
        if student_number.lower() == 'q':
            print("Exiting student data entry.")
            break

        if not student_number.isdigit():
            print("Invalid Student Number. Please enter a valid number.")
            continue

        student_data = [int(student_number), lesson_id] 
        for criterion in criteria:
            while True:
                score = input(f"Enter score for {criterion} (0-100): ").strip()
                if not score.isdigit() and score.lower() != 'q':
                    print("Invalid input. Please enter a valid numeric score.")
                    continue
                if score.lower() == 'q':
                    print("Exiting student data entry.")
                    conn.close()
                    return
                score = float(score)
                if 0 <= score <= 100:
                    student_data.append(score)
                    break
                else:
                    print("Please enter a valid score between 0 and 100.")

        # Students tablosuna verinin eklenmesi
        columns = ["Student", "lesson_id"] + [f"[{criterion}]" for criterion in criteria]  
        placeholders = ", ".join(["?"] * len(columns))
        insert_query = f"INSERT INTO Students ({', '.join(columns)}) VALUES ({placeholders});"

        try:
            cursor.execute(insert_query, student_data)
            conn.commit()
            print(f"Student data for Student ID {student_number} in Lesson ID {lesson_id} has been successfully added.")
        except Exception as e:
            print(f"An error occurred while inserting student data: {e}")
            conn.rollback()

    conn.close()


def insert_lesson():
    lesson_id = input("Enter the Lesson ID: ").strip()
    lesson_name = input("Enter the name of the new lesson: ").strip()

    conn = get_connection("RelationMatrix")
    conn.autocommit = True
    cursor = conn.cursor()

    cursor.execute('SELECT id FROM Lessons WHERE id = ?', lesson_id)
    existing_lesson = cursor.fetchone()

    if existing_lesson:
        print(f"Lesson with ID {lesson_id} already exists.")
    else:
        # Lessons tablosuna yeni ders ekleme sorgusu
        cursor.execute('''SET IDENTITY_INSERT Lessons ON;
            INSERT INTO Lessons (id, name) VALUES (?, ?);
            SET IDENTITY_INSERT Lessons OFF;
        ''', lesson_id, lesson_name)
        print(f"Lesson '{lesson_name}' with ID {lesson_id} has been added.")

    conn.close()


def menu():
    while True:
        print("\nSelect an action:")
        print("1. Add Lesson")
        print("For TABLE 1:")
        print("2. Add to Program Outcomes")
        print("3. Add to Course Outcomes")
        print("4. Add Relations")
        print("\nFor TABLE 2:")
        print("5. Add Evaluation Criteria")
        print("6. Add CourseOutcome-Criteria Relations")
        print("\nFor TABLE 4:")
        print("7. Add Student")
        print("\n8. Exit")

        choice = input("Enter your choice (1-8): ").strip()

        if choice in ['2', '3', '4', '5', '6','7']:
            lesson_id = input("Enter Lesson ID: ").strip()
        else:
            lesson_id = None  

        if choice == '1':
            insert_lesson() 
        elif choice == '2':
            get_input_and_insert_table('ProgramOutcomes', lesson_id)
        elif choice == '3':
            get_input_and_insert_table('CourseOutcomes', lesson_id)
        elif choice == '4':
            get_input_and_insert_relations(lesson_id)
        elif choice == '5':
            get_evaluation_criteria_and_insert(lesson_id)
        elif choice == '6':
            get_input_and_insert_evaluation_relations(lesson_id)
        elif choice == '7':
            create_students_table(lesson_id)
            add_student(lesson_id) 
        elif choice == '8':
            print("Exiting the program.")
            break
        else:
            print("Invalid choice. Please enter a number between 1 and 8.")

check_database()
check_tables()
# clear_relations()
menu()
create_table1()
create_table2()
create_table3()
save_table3_to_database()
create_table4()
create_notes()
save_table4_to_database()
create_table5()